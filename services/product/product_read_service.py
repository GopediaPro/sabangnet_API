import io
import os
import pandas as pd
import urllib.parse
import tempfile
from typing import List, Dict, Optional, Tuple, Any
from datetime import datetime
from fastapi.responses import StreamingResponse

from utils.logs.sabangnet_logger import get_logger
from utils.mappings.product_create_field_db_mapping import PRODUCT_CREATE_FIELD_MAPPING

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.inspection import inspect
from models.product.product_raw_data import ProductRawData
from repository.product_repository import ProductRepository
from schemas.product.product_raw_data_dto import ProductRawDataDto
from utils.exceptions.http_exceptions import DataNotFoundException
from utils.mappings.test_product_filed_db_mappring_for_excel import TEST_PRODUCT_DB_TO_EXCEL_HEADER, db_row_to_excel_row_generic

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

logger = get_logger(__name__)


class ProductReadService:
    def __init__(self, session: AsyncSession):
        self.session = session
        self.product_repository = ProductRepository(session)

    async def get_products_all(self, skip: int, limit: int) -> list[ProductRawDataDto]:
        products = await self.product_repository.get_products_all(skip=skip, limit=limit)
        dtos = [ProductRawDataDto.model_validate(product) for product in products]
        return dtos

    async def get_products_by_pagenation(self, page: int, page_size: int) -> list[ProductRawDataDto]:
        products = await self.product_repository.get_products_by_pagenation(page=page, page_size=page_size)
        dtos = [ProductRawDataDto.model_validate(product) for product in products]
        return dtos

    async def get_product_by_compayny_goods_cd(self, compayny_goods_cd: str) -> ProductRawData:
        return await self.product_repository.find_product_raw_data_by_company_goods_cd(compayny_goods_cd)
    
    async def get_product_by_product_nm_and_gubun(self, product_nm: str, gubun: str) -> ProductRawDataDto:
        res = await self.product_repository.find_product_raw_data_by_product_nm_and_gubun(product_nm, gubun)
        if res is None:
            raise ValueError(f"Product raw data not found: {product_nm}")
        return ProductRawDataDto.model_validate(res)
    
    async def get_product_raw_data_all(self) -> list[ProductRawDataDto]:
        objects = await self.product_repository.get_product_raw_data_all()
        return [ProductRawDataDto.model_validate(obj) for obj in objects]
    
    async def get_product_raw_data_by_company_goods_cds(self, company_goods_cds: List[str]) -> list[ProductRawDataDto]:
        """
        특정 company_goods_cd 목록으로 product_raw_data 조회
        Args:
            company_goods_cds: 조회할 company_goods_cd 목록
        Returns:
            ProductRawDataDto 목록
        """
        objects = await self.product_repository.get_product_raw_data_by_company_goods_cds(company_goods_cds)
        return [ProductRawDataDto.model_validate(obj) for obj in objects]
    
    async def get_product_raw_data_count(self) -> int:
        return await self.product_repository.count_product_raw_data()
    
    async def convert_product_data_to_excel(self) -> StreamingResponse:
            dto_dict_list = await self.get_product_raw_data_all()

            mapping_key_list = []
            mapping_value_list = []
            for k, v in PRODUCT_CREATE_FIELD_MAPPING.items():
                mapping_key_list.append(k)
                if isinstance(v, tuple):
                    for code in v:
                        mapping_value_list.append(code.lower())
                    continue
                mapping_value_list.append(v.lower())
            
            # class_cd1~4 등을 마이카테고리로 합쳐주는 로직
            category_join_list = []
            for dto_dict in dto_dict_list:
                for dto_k, dto_v in dto_dict.items():
                    if "class_cd" in dto_k and dto_v:
                        category_join_list.append(dto_v)
                my_category = ">".join(category_join_list)
                dto_dict["std_category"] = "" # 없던 필드 추가
                dto_dict["my_category"] = my_category
                for i in range(1, 5):
                    if f"class_cd{i}" in dto_dict:
                        del dto_dict[f"class_cd{i}"]
            
            # 엑셀 만들기
            df = pd.DataFrame(dto_dict_list, columns=mapping_value_list)
            df.columns = mapping_key_list # 한글 헤더로 바꿈
            stream = io.BytesIO()
            with pd.ExcelWriter(stream, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='품번코드대량등록툴')
            stream.seek(0)

            # 한글 파일명을 URL 인코딩
            filename = "디자인업무일지.xlsx"
            encoded_filename = urllib.parse.quote(filename, safe='')

            return StreamingResponse(
                stream,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
            )


    async def convert_test_product_data_to_excel_file_by_filter(
        self,
        sort_order: Optional[str] = None,
        created_before: Optional[datetime] = None,
        filename: str = "대량상품등록.xlsx",
        sheet_name: str = "품번코드대량등록툴",
    ) -> Tuple[str, str, int, int]:
        """
        대량상품등록(test_product_raw_data) form에 맞춰서 excel 파일 생성

        1) DB 조회 → dict 변환
        2) 각 행을 TEST_PRODUCT_DB_TO_EXCEL_HEADER 기준으로 매핑
        3) pandas로 저장 → openpyxl로 헤더 스타일/열 너비 적용
        4) 파일 경로/파일명/레코드 수/파일 크기(bytes) 반환

        Returns:
            (temp_file_path, file_name, record_count, file_size)
        """
        # 1) 데이터 조회
        rows = await self.product_repository.fetch_test_product_raw_data_for_excel(
            sort_order=sort_order,
            created_before=created_before,
        )
        dict_rows: List[Dict[str, Any]] = [self.product_repository.to_dict(r) for r in rows]
        if not dict_rows:
            raise DataNotFoundException("다운로드할 상품 데이터가 없습니다.")

        # 2) DB dict -> 엑셀 헤더 dict 변환 (매핑 기준으로 모든 헤더 보장, 값 없으면 "")
        headers: List[str] = list(TEST_PRODUCT_DB_TO_EXCEL_HEADER.values())
        mapped_rows: List[Dict[str, Any]] = []
        for db_row in dict_rows:
            excel_row = db_row_to_excel_row_generic(db_row, TEST_PRODUCT_DB_TO_EXCEL_HEADER)
            mapped_rows.append(excel_row)

        # 3) DataFrame 빌드(헤더 순서 강제)
        df = pd.DataFrame(mapped_rows)
        # 누락된 헤더가 있더라도 순서와 공란 유지되도록 reindex
        df = df.reindex(columns=headers)

        # 4) 임시 파일 생성 및 저장 (xlsxwriter 엔진으로 저장 → openpyxl 로드하여 스타일)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_path = tmp.name
        tmp.close()

        with pd.ExcelWriter(tmp_path, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        # 5) openpyxl로 다시 열어 헤더 스타일/열 너비/줄바꿈 적용
        wb = load_workbook(tmp_path)
        ws = wb[sheet_name]

        self._apply_header_style(ws, headers)
        self._auto_fit_columns(ws, headers)
        self._enable_wrap_text(ws)

        wb.save(tmp_path)

        # 6) 메타데이터 계산 및 반환
        record_count = len(df)
        file_size = os.path.getsize(tmp_path)

        logger.info(
            f"[convert_test_product_data_to_excel_file_by_filter] 생성 완료 - "
            f"file={filename}, count={record_count}, size={file_size}, path={tmp_path}"
        )
        return tmp_path, filename, record_count, file_size


    def _apply_header_style(self, ws, headers: List[str]) -> None:
        """
        헤더 행(1행)의 폰트/정렬 스타일을 적용합니다.
        """
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_fit_columns(self, ws, headers: List[str]) -> None:
        """
        각 열의 최대 문자열 길이를 계산하여 열 너비를 자동 조정합니다.
        """
        for col_idx, _ in enumerate(headers, 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for c in ws[letter]:
                if c.value is not None:
                    max_len = max(max_len, len(str(c.value)))
            ws.column_dimensions[letter].width = max_len + 2

    def _enable_wrap_text(self, ws) -> None:
        """
        데이터 셀에 줄바꿈을 활성화합니다(헤더 제외).
        """
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
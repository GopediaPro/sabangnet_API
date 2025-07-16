import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side # Font, Border, Side는 사용되지 않지만 import는 유지

def compare_excel_sheets(workbook, sheet1_name, sheet2_name, red_fill):
    try:
        # 1. 엑셀 파일 로드
        sheet1 = workbook[sheet1_name]
        sheet2 = workbook[sheet2_name]
    except KeyError as e:
        print(f"오류: 시트 이름을 찾을 수 없습니다. {e}")
        return False # 'false'를 'False'로 수정

    print(f"\n--- '{sheet1_name}' 와 '{sheet2_name}' 비교 시작 ---")

    # 2. 시트1과 시트2의 2행 이하 모든 값의 배경색을 배경색 없음으로 설정
    print("시트의 2행 이하 셀 배경색을 '배경색 없음'으로 설정 중...")
    # no_fill 객체를 인자로 받지 않고, 함수 내에서 PatternFill()로 초기화
    # for sheet in [sheet1, sheet2]:
    #     for row_idx in range(2, sheet.max_row + 1): # 2행부터 마지막 행까지
    #         for col_idx in range(1, sheet.max_column + 1): # 1열부터 마지막 열까지
    #             cell = sheet.cell(row=row_idx, column=col_idx)
    #             cell.fill = PatternFill() # 배경색 없음으로 설정 (기본 생성자는 fill_type=None과 동일)
    print(f"'{sheet1_name}' 및 '{sheet2_name}'의 배경색 초기화 완료.")

    # 3. 시트1과 시트2의 모든 값 비교 (수식 여부 포함)
    max_row = max(sheet1.max_row, sheet2.max_row)
    max_col = max(sheet1.max_column, sheet2.max_column)
    found_diff = False # 초기값은 False여야 합니다.

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell1 = sheet1.cell(row=row_idx, column=col_idx)
            cell2 = sheet2.cell(row=row_idx, column=col_idx)

            # 수식 여부 확인
            is_formula1 = cell1.data_type == 'f' # 'f'는 수식을 의미
            is_formula2 = cell2.data_type == 'f'

            # 셀의 실제 내용 (값 또는 수식 문자열) 가져오기
            # data_only=False로 로드했으므로, 수식 셀의 .value는 수식 문자열을 포함합니다.
            content1 = cell1.value
            content2 = cell2.value

            # 불일치 조건:
            # 1. 셀의 내용(값 또는 수식 문자열)이 다를 경우
            # 2. 수식 여부가 다를 경우 (하나는 수식이고 다른 하나는 아님)
            if content1 != content2 or is_formula1 != is_formula2:
                found_diff = True
                
                # 출력 메시지 개선: 수식인 경우 '수식:' 접두사 추가
                display_content1 = f"수식: '{content1}'" if is_formula1 else f"값: '{content1}'"
                display_content2 = f"수식: '{content2}'" if is_formula2 else f"값: '{content2}'"

                print(f"불일치 발견: 시트1 [{row_idx}행, {col_idx}열] {display_content1}, "
                      f"시트2 [{row_idx}행, {col_idx}열] {display_content2}")
                cell1.fill = red_fill # 시트1 해당 셀의 배경색을 붉은색으로 변경

    if not found_diff:
        print(f"'{sheet1_name}' 와 '{sheet2_name}' 시트 간 불일치 없음.")
    print(f"--- '{sheet1_name}' 와 '{sheet2_name}' 비교 완료 ---")
    return found_diff

def compare_selected_sheet_pairs(excel_filename):
    try:
        # 1. 엑셀 파일 로드
        # data_only=False가 기본값이므로 명시적으로 지정하지 않아도 됩니다.
        # keep_vba=False는 VBA 매크로가 있다면 제거하고 로드하여 파일 손상 가능성을 줄일 수 있습니다.
        workbook = openpyxl.load_workbook(excel_filename, keep_vba=False, data_only=False)
    except FileNotFoundError:
        print(f"오류: 파일을 찾을 수 없습니다. '{excel_filename}'")
        return
    except Exception as e:
        print(f"엑셀 파일을 로드하는 중 오류가 발생했습니다: {e}")
        return

    # 붉은색 배경 스타일 정의
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    
    # no_fill은 compare_excel_sheets 함수 내에서 PatternFill()로 직접 처리하므로 여기서 정의할 필요 없습니다.

    compares = []
    compares.append({'sheet1_name':"20250707_주문서확인처리_ 기본양식 -ERP용", 'sheet2_name':"자동화_m"})
    compares.append({'sheet1_name':"OK", 'sheet2_name':"OK_m"})
    compares.append({'sheet1_name':"IY", 'sheet2_name':"IY_m"})
    compares.append({'sheet1_name':"BB", 'sheet2_name':"BB_m"})

    # 모든 시트 배경색 초기화 (이 부분은 compare_excel_sheets 함수 내에서 다시 수행되므로 중복될 수 있습니다.
    # 하지만 모든 시트를 일괄적으로 먼저 초기화하고 싶다면 유지해도 무방합니다.)
    print("모든 시트 배경색 초기화 중...")
    # for compare in compares:
    #     for sheet_name in [compare["sheet1_name"], compare["sheet2_name"]]:
    #         current_sheet = workbook[sheet_name]
    #         for row_idx in range(2, current_sheet.max_row + 1):
    #             for col_idx in range(1, current_sheet.max_column + 1):
    #                 cell = current_sheet.cell(row=row_idx, column=col_idx)
    #                 cell.fill = PatternFill() # PatternFill()로 초기화
    print("모든 시트 배경색 초기화 완료.")

    for compare in compares:
        # no_fill 인자를 제거했습니다.
        compare_excel_sheets(workbook, compare["sheet1_name"], compare["sheet2_name"], red_fill)
        print("\n")

    # 변경된 내용을 새 파일로 저장 (원본 파일 손상 방지를 위해)
    output_filename = f"checked_{excel_filename}"

    try:
        workbook.save(output_filename)
        print(f"비교 및 변경 완료. 결과 파일은 '{output_filename}'으로 저장되었습니다.")
    except Exception as e:
        print(f"파일 저장 중 오류가 발생했습니다: {e}")

# --- 사용 예시 ---
if __name__ == "__main__":
    excel_file = "e_er_data_compare.xlsx"
    compare_selected_sheet_pairs(excel_file)
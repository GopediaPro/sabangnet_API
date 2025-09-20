from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


"""
test_product_raw_data (ORM: ProductRawData) 전용
DB 컬럼(영문, snake_case) -> 엑셀 헤더(한글)
"""

TEST_PRODUCT_DB_TO_EXCEL_HEADER: Dict[str, str] = {
    # 디자인업무일지 전용 + 기본 정보
    "no_product": "순번",
    "img_path_check": "대표이미지확인",             # DB에 없으면 "" 반환
    "detail_img_url": "상세이미지확인",
    "no_word": "글자수",
    "no_keyword": "키워드 Btyte",
    "product_nm": "모델명",
    "gubun": "구분",

    # 기본 상품 정보
    "goods_nm": "상품명 [필수]",
    "goods_keyword": "상품약어",
    "model_nm": "모델명",
    "model_no": "모델NO",
    "brand_nm": "브랜드명",
    "compayny_goods_cd": "자체상품코드",   # ORM 오타(compayny) 유지
    "goods_search": "사이트검색어",
    "std_category": "표준카테고리",
    "goods_gubun": "상품구분 [필수]",
    "my_category": "마이카테고리",
    "partner_id": "매입처ID",
    "dpartner_id": "물류처ID",
    "maker": "제조사 [필수]",
    "origin": "원산지(제조국) [필수]",
    "make_year": "생산연도",
    "make_dm": "제조일",
    "goods_season": "시즌",
    "sex": "남녀구분",
    "status": "상품상태 [필수]",
    "deliv_able_region": "판매지역",
    "tax_yn": "세금구분 [필수]",
    "delv_type": "배송비구분 [필수]",
    "delv_cost": "배송비",
    "banpum_area": "반품지구분",
    "goods_cost": "원가 [필수]",
    "goods_price": "판매가 [필수]",
    "goods_consumer_price": "TAG가 [필수]",

    # 옵션
    "char_1_nm": "옵션제목(1)",
    "char_1_val": "옵션상세명칭(1)",
    "char_2_nm": "옵션제목(2)",
    "char_2_val": "옵션상세명칭(2)",

    # 이미지
    "img_path": "대표이미지 [필수]",
    "img_path1": "종합몰(JPG)이미지",
    "img_path2": "부가이미지2",
    "img_path3": "부가이미지3",
    "img_path4": "부가이미지4",
    "img_path5": "부가이미지5",
    "img_path6": "부가이미지6",
    "img_path7": "부가이미지7",
    "img_path8": "부가이미지8",
    "img_path9": "부가이미지9",
    "img_path10": "부가이미지10",

    # 상세/인증
    "goods_remarks": "상품상세설명 [필수]",
    "pack_code_str": "추가상품그룹코드",
    "certno": "인증번호",
    "avlst_dm": "인증유효시작일",
    "avled_dm": "인증유효마지막일",
    "issuedate": "발급일자",
    "certdate": "인증일자",
    "cert_agency": "인증기관",
    "certfield": "인증분야",
    "stock_use_yn": "재고관리사용여부",
    "expire_dm": "유효일",
    "material": "식품 재료/원산지",
    "goods_cost2": "원가2",

    # 부가 이미지
    "img_path11": "부가이미지11 (Cafe24, MakeShop/고도몰/패션플러스 리스트이미지)",
    "img_path12": "부가이미지12",
    "img_path13": "부가이미지13",
    "supply_save_yn": "합포시 제외 여부",
    "img_path14": "부가이미지14",
    "img_path15": "부가이미지15",
    "img_path16": "부가이미지16",
    "img_path17": "부가이미지17",
    "img_path18": "부가이미지18",
    "img_path19": "부가이미지19",
    "img_path20": "부가이미지20",
    "img_path21": "부가이미지21",
    "img_path22": "부가이미지22",

    # 기타
    "descrition": "관리자메모",   # DB 오타(descrition) 유지
    "opt_type": "옵션수정여부",
    "goods_nm_en": "영문 상품명",
    "goods_nm_pr": "출력 상품명",
    "img_path23": "인증서이미지",
    "goods_remarks2": "추가 상품상세설명_1",
    "goods_remarks3": "추가 상품상세설명_2",
    "goods_remarks4": "추가 상품상세설명_3",
    "origin2": "원산지 상세지역",
    "importno": "수입신고번호",
    "img_path24": "수입면장이미지",

    # 속성
    "prop1_cd": "속성분류코드 [필수]",
    "prop_val1": "속성값1 [필수/선택]",
    "prop_val2": "속성값2 [필수/선택]",
    "prop_val3": "속성값3 [필수/선택]",
    "prop_val4": "속성값4 [필수/선택]",
    "prop_val5": "속성값5 [필수/선택]",
    "prop_val6": "속성값6 [필수/선택]",
    "prop_val7": "속성값7 [필수/선택]",
    "prop_val8": "속성값8 [필수/선택]",
    "prop_val9": "속성값9 [필수/선택]",
    "prop_val10": "속성값10 [필수/선택]",
    "prop_val11": "속성값11 [필수/선택]",
    "prop_val12": "속성값12 [필수/선택]",
    "prop_val13": "속성값13 [필수/선택]",
    "prop_val14": "속성값14 [필수/선택]",
    "prop_val15": "속성값15 [필수/선택]",
    "prop_val16": "속성값16 [필수/선택]",
    "prop_val17": "속성값17 [필수/선택]",
    "prop_val18": "속성값18 [필수/선택]",
    "prop_val19": "속성값19 [필수/선택]",
    "prop_val20": "속성값20 [필수/선택]",
    "prop_val21": "속성값21 [필수/선택]",
    "prop_val22": "속성값22 [필수/선택]",
    "prop_val23": "속성값23 [필수/선택]",
    "prop_val24": "속성값24 [필수/선택]",
    "prop_val25": "속성값25 [필수/선택]",
    "prop_val26": "속성값26 [필수/선택]",
    "prop_val27": "속성값27 [필수/선택]",
    "prop_val28": "속성값28 [필수/선택]",
    "prop_val29": "속성값29 [필수/선택]",
    "prop_val30": "속성값30 [필수/선택]",
    "prop_val31": "속성값31 [필수/선택]",
    "prop_val32": "속성값32 [필수/선택]",
    "prop_val33": "속성값33 [필수/선택]",
    "prop_val34": "속성값34 [필수/선택]",
    "prop_val35": "속성값35 [필수/선택]",
    "prop_val36": "속성값36 [필수/선택]",
    "prop_val37": "속성값37 [필수/선택]",
    "prop_val38": "속성값38 [필수/선택]",
}


def db_row_to_excel_row_generic(
    db_row: Dict[str, Any],
    mapping: Dict[str, str],
) -> Dict[str, Any]:
    """
    DB dict(영문 snake_case) -> 엑셀 헤더 dict(한글)
    - 매핑된 헤더는 무조건 생성, 값이 없으면 ""(빈 문자열)
    - None도 ""로 변환
    """
    out: Dict[str, Any] = {}
    lower_row = {k.lower(): v for k, v in db_row.items()}
    for eng, kor in mapping.items():
        val = lower_row.get(eng.lower(), "")
        out[kor] = "" if val is None else val
    return out


def export_to_excel_generic(
    db_rows: List[Dict[str, Any]],
    file_path: str,
    mapping: Dict[str, str],
    sheet_name: str = "시트1",
):
    """
    매핑을 받아 공통으로 엑셀을 내보내는 간단 유틸
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = list(mapping.values())
    ws.append(headers)

    for row in db_rows:
        excel_row = db_row_to_excel_row_generic(row, mapping)
        ws.append([excel_row.get(h, "") for h in headers])

    # 헤더 스타일 및 열 너비
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        max_len = 0
        for c in ws[get_column_letter(col_idx)]:
            if c.value is not None:
                max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    wb.save(file_path)

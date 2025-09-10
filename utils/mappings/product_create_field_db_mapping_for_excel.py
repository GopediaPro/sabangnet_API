from typing import Dict, Any
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font


# DB 컬럼(영문) -> 엑셀 헤더(한글)
PRODUCT_CREATE_DB_TO_EXCEL_HEADER: Dict[str, str] = {
    # 담당/코드/플래그
    "manager_nm": "담당자",                                # TODO(DB 없음 추정)
    "sabangnet_master_code": "사방넷품번코드(마스터)",     # TODO(DB 없음 추정)
    "sabangnet_special_code": "사방넷품번코드(전문몰)",    # TODO(DB 없음 추정)
    "one_plus_one_flag": "1+1",                           # TODO(DB 없음 추정)
    "vendor_nm": "입고업체",                               # TODO(DB 없음 추정)
    "collection_flag": "모음전",                           # TODO(DB 없음 추정)
    "soldout_reg_yn": "등록 품절여부",                     # TODO(DB 없음 추정)

    # 카테고리
    "class_nm1": "대분류",
    "class_nm4": "세분류",

    # 상품 기본
    "product_nm": "제품명",
    "goods_nm": "상품명",
    "detail_path_img": "상세페이지경로(이미지폴더)",

    # 배송/가격/검색
    "delv_cost": "배송비",
    "goods_search": "키워드",
    "goods_price": "판매가(유료배송)",

    # 인증/진행 옵션
    "certno": "인증번호",
    "char_process": "진행옵션 가져오기",

    # 옵션
    "char_1_nm": "옵션명1",
    "char_1_val": "옵션상세1",
    "char_2_nm": "옵션명2",
    "char_2_val": "옵션상세2",

    # 이미지
    "img_path":  "대표이미지",
    "img_path1": "부가이미지1",
    "img_path2": "부가이미지2",
    "img_path3": "부가이미지3",
    "img_path4": "부가이미지4",
    "img_path5": "부가이미지5",

    # 상세/배너/URL/1+1 옵션
    "goods_remarks":        "상세설명",
    "mobile_bn":            "모바일배너",
    "one_plus_one_bn":      "1+1배너",
    "goods_remarks_url":    "상세설명url",
    "delv_one_plus_one":    "1+1옵션",
    "delv_one_plus_one_detail":    "",
}


def db_row_to_excel_row(db_row: Dict[str, Any],
                        mapping: Dict[str, str] = PRODUCT_CREATE_DB_TO_EXCEL_HEADER) -> Dict[str, Any]:
    """
    DB dict -> 엑셀 헤더 dict 변환
    - 매핑된 키만 변환
    - 매핑에 없는 값은 무시
    - 매핑된 헤더는 무조건 생성 (DB에 없으면 "")
    """
    result: Dict[str, Any] = {}
    lower_row = {k.lower(): v for k, v in db_row.items()}
    for eng, kor in mapping.items():
        val = lower_row.get(eng.lower(), "")
        result[kor] = "" if val is None else val
    return result


def export_to_excel(db_rows: list[Dict[str, Any]], file_path: str,
                    mapping: Dict[str, str] = PRODUCT_CREATE_DB_TO_EXCEL_HEADER):
    """
    DB dict 리스트를 받아서 엑셀 파일로 내보내기
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "상품등록"

    # 헤더 작성
    headers = list(mapping.values())
    ws.append(headers)

    # 데이터 작성
    for row in db_rows:
        excel_row = db_row_to_excel_row(row, mapping)
        ws.append([excel_row.get(header, "") for header in headers])

    # 스타일 (자동 줄바꿈, bold header, 열 너비 조정)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 열 너비 자동 조정 (모든 값 str 변환 후 길이 체크)
        max_length = 0
        for c in ws[get_column_letter(col_idx)]:
            if c.value is not None:
                max_length = max(max_length, len(str(c.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    wb.save(file_path)

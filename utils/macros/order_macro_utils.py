# std
import re
from collections import defaultdict
import openpyxl
from openpyxl import load_workbook
# utils
from utils.logs.sabangnet_logger import get_logger
# erp
from utils.macros.ERP.ali_erp_macro import ERPAliMacro
from utils.macros.ERP.g_a_erp_macro import ERPGmaAucMacro
from utils.macros.ERP.etc_site_macro import ERPEtcSiteMacro
from utils.macros.ERP.zigzag_erp_macro import ERPZigzagMacro
from utils.macros.ERP.brandi_erp_macro import ERPBrandiMacro
# dto
from schemas.down_form_orders.down_form_order_dto import DownFormOrderDto
# bundle
from utils.macros.happojang.gok_merge_packaging import gok_merge_packaging
from utils.macros.happojang.ali_merge_packaging import ali_merge_packaging
from utils.macros.happojang.zigzag_merge_packaging import zigzag_merge_packaging
from utils.macros.happojang.brandy_merge_packaging import brandy_merge_packaging
from utils.macros.happojang.etc_site_merge_packaging import etc_site_merge_packaging


logger = get_logger(__name__)


class OrderMacroUtils:
    """
    매크로 실행 관련 유틸리티 클래스
    """

    def __init__(self):
        # macro_name과 실제 실행 함수를 매핑
        self.MACRO_MAP = {
            "AliMacro": self.run_ali_macro,
            "ZigzagMacro": self.run_zigzag_macro,
            "BrandiMacro": self.run_brandi_macro,
            "ECTSiteMacro": self.run_etc_site_macro,
            "GmarketAuctionMacro": self.run_gmarket_auction_macro,
            "ali_merge_packaging": ali_merge_packaging,
            "gok_merge_packaging": gok_merge_packaging,
            "zigzag_merge_packaging": zigzag_merge_packaging,
            "brandy_merge_packaging": brandy_merge_packaging,
            "etc_site_merge_packaging": etc_site_merge_packaging,
        }

    def run_ali_macro(self, file_path: str) -> int:
        return ERPAliMacro(file_path).ali_erp_macro_run()

    def run_etc_site_macro(self, file_path: str) -> int:
        return ERPEtcSiteMacro(file_path).etc_site_macro_run()

    def run_gmarket_auction_macro(self, file_path: str) -> int:
        return ERPGmaAucMacro(file_path).gauc_erp_macro_run()

    def run_brandi_macro(self, file_path: str) -> int:
        return ERPBrandiMacro(file_path).brandi_erp_macro_run()

    def run_zigzag_macro(self, file_path: str) -> int:
        return ERPZigzagMacro(file_path).zigzag_erp_macro_run()

    def modify_site_column_for_star_delivery(self, file_path: str) -> str:
        """
        is_star=True일 때 B열(사이트) 값에 "-스타배송" 추가
        Args:
            file_path: Excel 파일 경로
        Returns:
            수정된 파일 경로
        """
        try:
            # Excel 파일 로드
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            
            # B열(사이트)의 모든 셀을 순회하며 수정
            for row in range(2, worksheet.max_row + 1):  # 헤더 제외하고 2행부터 시작
                cell = worksheet[f'B{row}']
                if cell.value:
                    original_value = str(cell.value)
                    modified_value = self._add_star_delivery_suffix(original_value)
                    if modified_value != original_value:
                        cell.value = modified_value
                        logger.info(f"B{row} 수정: '{original_value}' -> '{modified_value}'")
            
            # 수정된 파일 저장
            workbook.save(file_path)
            logger.info(f"스타배송 수정 완료: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"스타배송 수정 중 오류 발생: {e}")
            raise

    def _add_star_delivery_suffix(self, site_value: str) -> str:
        """
        사이트 값에 "-스타배송" 추가
        Args:
            site_value: 원본 사이트 값 (예: "[베이지베이글]G마켓2.0")
        Returns:
            수정된 사이트 값 (예: "[베이지베이글-스타배송]G마켓2.0")
        """
        if not site_value:
            return site_value
        
        # ']' 문자를 찾아서 그 앞에 "-스타배송" 추가
        if ']' in site_value:
            # ']' 위치를 찾아서 그 앞에 "-스타배송" 삽입
            bracket_end_index = site_value.find(']')
            if bracket_end_index > 0:
                # ']' 앞에 이미 "-스타배송"이 있는지 확인
                before_bracket = site_value[:bracket_end_index]
                if "-스타배송" not in before_bracket:
                    modified_value = site_value[:bracket_end_index] + "-스타배송" + site_value[bracket_end_index:]
                    return modified_value
        
        return site_value

    def process_orders_for_db(self, dto_list: list[DownFormOrderDto]) -> list[dict]:
        """DB 저장을 위한 주문 리스트 전체 처리"""

        order_list = [order.model_dump() for order in dto_list]
        
        # 1. 개별 주문 전처리
        processed_orders = [self.preprocess_order_data(order) for order in order_list]
        
        # 2. 정렬
        processed_orders = self.sort_orders_by_receiver_and_site(processed_orders)
        
        # 3. 중복 제거 (필요한 경우)
        if processed_orders['site_name'] in ['G마켓', '옥션']:
            processed_orders = self.remove_duplicate_shipping_fee(processed_orders)
        
        # 4. 주문 병합 (브랜디의 경우)
        if '브랜디' in processed_orders['site_name']:
            groups = self.group_orders_by_product_and_receiver(processed_orders)
            processed_orders = [self.merge_order_group(group) for group in groups.values()]
        
        return processed_orders

    def preprocess_order_data(self, order_data: dict) -> dict:
        """주문 데이터 전체 전처리"""
        
        # 기본 정보 정리
        processed = {
            'receiver_name': str(order_data.get('receiver_name', '')).strip(),
            'item_name': self.clean_item_name(order_data.get('item_name')),
            'phone1': self.format_phone_number(order_data.get('phone1')),
            'phone2': self.format_phone_number(order_data.get('phone2')),
            'address': str(order_data.get('address', '')).strip(),
            'order_number': self.clean_order_text(order_data.get('order_number')),
        }
        
        # 숫자 데이터 변환
        numeric_fields = ['quantity', 'product_price', 'option_price', 'delv_cost', 'total_amount']
        for field in numeric_fields:
            processed[field] = self.convert_to_number(order_data.get(field, 0))
        
        # 배송비 계산
        processed['delv_cost'] = self.calculate_delivery_fee(processed)
        
        # 제주 주소 처리
        if self.is_jeju_address(processed['address']):
            processed['item_name'] = self.add_jeju_notice(processed['item_name'], processed['address'])
            processed['is_jeju'] = True
        else:
            processed['is_jeju'] = False
        
        # 다중 수량 체크
        processed['is_multiple_quantity'] = self.check_multiple_quantities(processed['item_name'])
        
        return processed

    def sum_slash_values(self, value_str: str) -> float:
        """슬래시로 구분된 값들을 합산"""

        if not value_str or '/' not in str(value_str):
            return self.convert_to_number(value_str)
        
        total = 0.0
        for part in str(value_str).split('/'):
            total += self.convert_to_number(part.strip())
        
        return total

    def get_first_valid_slash_value(self, value_str: str) -> float:
        """슬래시로 구분된 값 중 첫 번째 유효 숫자만 반환"""

        if not value_str or '/' not in str(value_str):
            return self.convert_to_number(value_str)
        
        for part in str(value_str).split('/'):
            num = self.convert_to_number(part.strip())
            if num != 0:
                return num
        
        return 0.0

    def is_jeju_address(self, address: str) -> bool:
        """주소가 제주도인지 확인"""

        if not address:
            return False
        
        jeju_keywords = ['제주', '서귀포']
        return any(keyword in str(address) for keyword in jeju_keywords)

    def add_jeju_notice(self, item_name: str, address: str) -> str:
        """제주 주소인 경우 상품명에 안내문구 추가"""

        if not self.is_jeju_address(address):
            return item_name
        
        notice = "[3000원 연락해야함]"
        if notice not in str(item_name):
            return f"{item_name} {notice}"
        
        return item_name

    def group_orders_by_product_and_receiver(self, orders: list[dict]) -> dict:
        """상품번호+수령인 기준으로 주문 그룹핑"""

        groups = defaultdict(list)
        
        for order in orders:
            key = f"{order.get('product_code', '')}-{order.get('receiver_name', '')}"
            groups[key].append(order)
        
        return groups

    def merge_order_group(self, order_group: list[dict]) -> dict:
        """같은 그룹의 주문들을 병합"""

        if len(order_group) <= 1:
            return order_group[0] if order_group else {}
        
        base_order = order_group[0].copy()
        
        # D열(금액) 합산
        total_amount = 0.0
        for order in order_group:
            amount = order.get('total_amount', 0)
            if isinstance(amount, str) and '=' in amount:
                # 수식인 경우 개별 항목 합산
                o_val = self.convert_to_number(order.get('product_price', 0))
                p_val = self.convert_to_number(order.get('option_price', 0))
                v_val = self.convert_to_number(order.get('delv_cost', 0))
                total_amount += (o_val + p_val + v_val)
            else:
                total_amount += self.convert_to_number(amount)
        
        # G열(수량) 합산
        total_quantity = sum(self.convert_to_number(order.get('quantity', 0)) for order in order_group)
        
        # F열(item_name) 결합
        model_names = []
        for order in order_group:
            model = self.clean_item_name(order.get('item_name', ''))
            if model:
                model_names.append(model)
        
        base_order.update({
            'total_amount': total_amount,
            'quantity': total_quantity,
            'item_name': ' + '.join(model_names)
        })
        
        return base_order

    def calculate_delivery_fee(self, order_data: dict) -> int:
        """사이트별 배송비 계산"""
        
        # 무료 배송 사이트
        free_delivery_sites = {'오늘의집'}
        site_name = order_data.get('site_name', '')
        if any(site in site_name for site in free_delivery_sites):
            return 0
        
        # 배송비 분할 사이트
        split_delivery_sites = {'롯데온', '보리보리', '스마트스토어', '톡스토어'}
        if any(site in site_name for site in split_delivery_sites):
            order_code = str(order_data.get('idx', ''))
            original_fee = self.convert_to_number(order_data.get('delv_cost', 0))
            
            if '/' in order_code:
                order_count = len(order_code.split('/'))
                if original_fee > 3000 and order_count > 0:
                    return round(original_fee / order_count)
        
        # 토스 특수 처리
        if '토스' in site_name:
            product_amount = self.convert_to_number(order_data.get('pay_cost', 0))
            return 0 if product_amount > 30000 else 3000
        
        # 기본 배송비 반환
        return self.convert_to_number(order_data.get('delv_cost', 0))

    def sort_orders_by_receiver_and_site(self, orders: list[dict]) -> list[dict]:
        """수취인명, 사이트 순으로 정렬"""
        return sorted(orders, key=lambda x: (
            x.get('receiver_name', ''),
            x.get('site_name', '')
        ))

    def sort_orders_by_receiver_only(self, orders: list[dict]) -> list[dict]:
        """수취인명으로만 정렬"""
        return sorted(orders, key=lambda x: x.get('receiver_name', ''))

    def build_lookup_mapping(self, lookup_data: list[dict]) -> dict:
        """조회 테이블을 딕셔너리로 변환"""
        mapping = {}
        for row in lookup_data:
            key = str(row.get('lookup_key', ''))
            value = row.get('lookup_value', 'S')  # 기본값 'S'
            if key:
                mapping[key] = value
        return mapping

    def lookup_value(self, lookup_map: dict, key: str, default: str = 'S') -> str:
        """키로 값 조회"""

        return lookup_map.get(str(key), default)

    def remove_duplicate_shipping_fee(self, orders: list[dict]) -> list[dict]:
        """바구니 번호별 중복 배송비 제거"""

        basket_dict = {}
        
        # 첫 번째 패스: 배송비가 있는 첫 주문 기록
        for i, order in enumerate(orders):
            basket_no = str(order.get('mall_order_id', '')).strip()
            delv_cost = self.convert_to_number(order.get('delv_cost', 0))
            
            if basket_no and basket_no != '':
                if basket_no not in basket_dict and delv_cost > 0:
                    basket_dict[basket_no] = i
        
        # 두 번째 패스: 중복 바구니의 배송비를 0으로 설정
        for i, order in enumerate(orders):
            basket_no = str(order.get('mall_order_id', '')).strip()
            
            if basket_no and basket_no != '':
                if basket_no in basket_dict and basket_dict[basket_no] != i:
                    order['delv_cost'] = 0
        
        return orders

    def check_multiple_quantities(self, product_text: str) -> bool:
        """상품명에서 다중 수량 여부 확인"""

        if not product_text:
            return False
        
        parts = [p.strip() for p in str(product_text).split('+')]
        qty_count = sum(1 for p in parts if re.search(r'\d+개', p))
        
        return qty_count >= 2

    def extract_quantity_from_text(self, text: str) -> int:
        """텍스트에서 수량 추출"""

        if not text:
            return 1
        
        match = re.search(r'(\d+)개', str(text))
        return int(match.group(1)) if match else 1

    def clean_item_name(self, val: str) -> str:
        """
        상품명에서 ' 1개' 등 불필요한 텍스트 제거
        """

        if not val:
            return val
        return str(val).replace(' 1개', '').strip()

    def format_phone_number(self, val: str) -> str:
        """
        전화번호 11자리 → 010-0000-0000 형식
        """

        val = str(val or '').replace('-', '').strip()
        if len(val) == 11 and val.startswith('010') and val.isdigit():
            return f"{val[:3]}-{val[3:7]}-{val[7:]}"
        return val

    def convert_to_number(self, val: str) -> int:
        """
        '12,345원' → 12345 (실패 시 0)
        """

        import re
        try:
            return int(re.sub(r"[^\d.-]", "", str(val))) if str(val).strip() else 0
        except ValueError:
            return 0

    def clean_order_text(self, val: str) -> str:
        """
        주문번호 등에서 불필요한 공백/특수문자 제거
        """
        
        if not val:
            return ''
        return str(val).strip().replace(' ', '').replace('\n', '').replace('\r', '')

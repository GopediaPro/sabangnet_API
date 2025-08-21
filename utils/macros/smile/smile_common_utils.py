import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Tuple, Any
from utils.logs.sabangnet_logger import get_logger
from utils.handlers.data_type_handler import DataTypeHandler

logger = get_logger(__name__)


class SmileCommonUtils:
    """
    스마일배송 매크로 공통 유틸리티 클래스
    보일러플레이트 코드를 공통 메서드로 분리
    """
    
    @staticmethod
    def apply_basic_formatting(ws, font_size: int = 9, row_height: int = 15):
        """
        기본 서식 적용 (폰트 크기, 행 높이)
        
        Args:
            ws: 워크시트
            font_size: 폰트 크기
            row_height: 행 높이
        """
        try:
            # 폰트 크기 설정
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(size=font_size)
            
            # 행 높이 설정
            for row in ws.iter_rows():
                for cell in row:
                    ws.row_dimensions[cell.row].height = row_height
                    
            logger.info("기본 서식 적용 완료")
            
        except Exception as e:
            logger.error(f"기본 서식 적용 중 오류: {str(e)}")
            raise
    
    @staticmethod
    def delete_colored_rows(ws):
        """
        색이 있는 행 삭제
        
        Args:
            ws: 워크시트
        """
        try:
            rows_to_delete = []
            last_row = ws.max_row
            last_col = ws.max_column
            
            for row_num in range(2, last_row + 1):
                is_colored = False
                for col_num in range(1, last_col + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    if (cell.fill.start_color.rgb != '00000000' and 
                        cell.fill.start_color.rgb != '00FFFFFF'):
                        is_colored = True
                        break
                if is_colored:
                    rows_to_delete.append(row_num)
            
            # 역순으로 삭제 (인덱스 변화 방지)
            for row_num in reversed(rows_to_delete):
                ws.delete_rows(row_num)
                
            logger.info(f"색이 있는 행 {len(rows_to_delete)}개 삭제 완료")
            
        except Exception as e:
            logger.error(f"색이 있는 행 삭제 중 오류: {str(e)}")
            raise
    
    @staticmethod
    def insert_columns(ws, start_col: int, num_cols: int, headers: List[str] = None):
        """
        열 삽입
        
        Args:
            ws: 워크시트
            start_col: 시작 열 번호
            num_cols: 삽입할 열 수
            headers: 헤더 리스트
        """
        try:
            ws.insert_cols(start_col, num_cols)
            
            if headers:
                for i, header in enumerate(headers):
                    col_letter = get_column_letter(start_col + i)
                    ws[f'{col_letter}1'] = header
                    
            logger.info(f"열 {num_cols}개 삽입 완료 (시작: {start_col})")
            
        except Exception as e:
            logger.error(f"열 삽입 중 오류: {str(e)}")
            raise
    
    @staticmethod
    def get_worksheet_headers(ws, max_cols: int = None) -> Dict[str, str]:
        """
        워크시트의 헤더 정보를 반환
        
        Args:
            ws: 워크시트
            max_cols: 확인할 최대 열 수 (None이면 모든 열 확인)
            
        Returns:
            Dict[str, str]: {열_문자: 헤더_값} 형태의 딕셔너리
        """
        try:
            headers = {}
            
            # 최대 열 수 결정
            if max_cols is None:
                max_cols = ws.max_column
            
            # 1행의 각 열에서 헤더 값 가져오기
            for col_num in range(1, max_cols + 1):
                col_letter = get_column_letter(col_num)
                cell_value = ws[f'{col_letter}1'].value
                headers[col_letter] = cell_value if cell_value is not None else ""
            
            logger.info(f"워크시트 헤더 정보 조회 완료 (총 {len(headers)}개 열)")
            return headers
            
        except Exception as e:
            logger.error(f"워크시트 헤더 조회 중 오류: {str(e)}")
            raise
    
    @staticmethod
    def print_worksheet_headers(ws, max_cols: int = None, show_empty: bool = False):
        """
        워크시트의 헤더 정보를 콘솔에 출력
        
        Args:
            ws: 워크시트
            max_cols: 확인할 최대 열 수 (None이면 모든 열 확인)
            show_empty: 빈 헤더도 출력할지 여부
        """
        try:
            headers = SmileCommonUtils.get_worksheet_headers(ws, max_cols)
            
            print(f"\n=== 워크시트 헤더 정보 ===")
            print(f"총 열 수: {len(headers)}")
            print("-" * 50)
            
            for col_letter, header_value in headers.items():
                if show_empty or header_value:
                    print(f"{col_letter}: {header_value}")
            
            print("-" * 50)
            
        except Exception as e:
            logger.error(f"워크시트 헤더 출력 중 오류: {str(e)}")
            raise
    
    @staticmethod
    def delete_columns(ws, start_col: int, num_cols: int):
        """
        열 삭제
        
        Args:
            ws: 워크시트
            start_col: 시작 열 번호
            num_cols: 삭제할 열 수
        """
        try:
            ws.delete_cols(start_col, num_cols)
            logger.info(f"열 {num_cols}개 삭제 완료 (시작: {start_col})")
            
        except Exception as e:
            logger.error(f"열 삭제 중 오류: {str(e)}")
            raise
    
    @staticmethod
    def set_cell_background(ws, cell_address: str, color: str = "FFFF00"):
        """
        셀 배경색 설정
        
        Args:
            ws: 워크시트
            cell_address: 셀 주소 (예: "A1")
            color: 색상 코드 (예: "FFFF00" - 노란색)
        """
        try:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws[cell_address].fill = fill
            logger.info(f"셀 {cell_address} 배경색 설정 완료")
            
        except Exception as e:
            logger.error(f"셀 배경색 설정 중 오류: {str(e)}")
            raise
    
    @staticmethod
    def copy_column_values(ws, source_col: str, target_col: str, start_row: int = 1):
        """
        열 값 복사
        
        Args:
            ws: 워크시트
            source_col: 원본 열
            target_col: 대상 열
            start_row: 시작 행 번호
        """
        try:
            last_row = ws.max_row
            for row_num in range(start_row, last_row + 1):
                source_cell = ws[f'{source_col}{row_num}']
                target_cell = ws[f'{target_col}{row_num}']
                target_cell.value = source_cell.value
                
            logger.info(f"열 복사 완료: {source_col} → {target_col}")
            
        except Exception as e:
            logger.error(f"열 복사 중 오류: {str(e)}")
            raise
    
    @staticmethod
    def apply_auto_filter(ws):
        """
        자동 필터 적용
        
        Args:
            ws: 워크시트
        """
        try:
            last_col = ws.max_column
            last_row = ws.max_row
            ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
            logger.info("자동 필터 적용 완료")
            
        except Exception as e:
            logger.error(f"자동 필터 적용 중 오류: {str(e)}")
            raise
    
    @staticmethod
    def sort_dataframe(df: pd.DataFrame, sort_columns: List[str], ascending: List[bool] = None):
        """
        DataFrame 정렬
        
        Args:
            df: DataFrame
            sort_columns: 정렬할 컬럼 리스트
            ascending: 오름차순 여부 리스트
            
        Returns:
            pd.DataFrame: 정렬된 DataFrame
        """
        try:
            if ascending is None:
                ascending = [True] * len(sort_columns)
                
            df_sorted = df.sort_values(by=sort_columns, ascending=ascending)
            logger.info(f"DataFrame 정렬 완료: {sort_columns}")
            return df_sorted
            
        except Exception as e:
            logger.error(f"DataFrame 정렬 중 오류: {str(e)}")
            raise
    
    @staticmethod
    def find_value_in_dataframe(df: pd.DataFrame, search_column: str, search_value: str, return_column: str) -> str:
        """
        DataFrame에서 값 찾기
        
        Args:
            df: DataFrame
            search_column: 검색할 컬럼
            search_value: 검색할 값
            return_column: 반환할 컬럼
            
        Returns:
            str: 찾은 값 또는 빈 문자열
        """
        try:
            if df.empty:
                return ""
                
            match = df[df[search_column] == search_value]
            if not match.empty:
                return str(match.iloc[0][return_column])
            return ""
            
        except Exception as e:
            logger.error(f"DataFrame 값 검색 중 오류: {str(e)}")
            return ""
    
    @staticmethod
    def convert_to_dataframe(data: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        딕셔너리 리스트를 DataFrame으로 변환
        
        Args:
            data: 딕셔너리 리스트
            
        Returns:
            pd.DataFrame: 변환된 DataFrame
        """
        try:
            if not data:
                return pd.DataFrame()
            
            df = pd.DataFrame(data)
            logger.info(f"DataFrame 변환 완료: {len(df)} 행")
            return df
            
        except Exception as e:
            logger.error(f"DataFrame 변환 중 오류: {str(e)}")
            return pd.DataFrame()
    
    @staticmethod
    def validate_required_fields(data: List[Dict[str, Any]], required_fields: List[str]) -> bool:
        """
        필수 필드 유효성 검증
        
        Args:
            data: 검증할 데이터 리스트
            required_fields: 필수 필드 리스트
            
        Returns:
            bool: 유효성 검증 결과
        """
        try:
            if not data:
                return False
            
            for item in data:
                if not all(field in item for field in required_fields):
                    return False
            
            logger.info(f"필수 필드 유효성 검증 완료: {len(data)} 항목")
            return True
            
        except Exception as e:
            logger.error(f"필수 필드 유효성 검증 중 오류: {str(e)}")
            return False
    
    @staticmethod
    def split_text_by_delimiter(text: str, delimiter: str = " ") -> List[str]:
        """
        텍스트를 구분자로 분할
        
        Args:
            text: 분할할 텍스트
            delimiter: 구분자
            
        Returns:
            List[str]: 분할된 텍스트 리스트
        """
        try:
            if not text:
                return []
            
            return text.split(delimiter)
            
        except Exception as e:
            logger.error(f"텍스트 분할 중 오류: {str(e)}")
            return []
    
    @staticmethod
    def create_sku_dictionary(sku_data: pd.DataFrame) -> Dict[str, str]:
        """
        SKU 데이터를 딕셔너리로 변환
        
        Args:
            sku_data: SKU DataFrame (sku_number, model_name 컬럼 포함)
            
        Returns:
            Dict[str, str]: SKU 딕셔너리 (sku_number -> model_name)
        """
        try:
            sku_dict = {}
            
            # 컬럼명이 있는 경우
            if 'sku_number' in sku_data.columns and 'model_name' in sku_data.columns:
                for _, row in sku_data.iterrows():
                    sku_number = str(row['sku_number']).strip() if pd.notna(row['sku_number']) else ""
                    model_name = str(row['model_name']).strip() if pd.notna(row['model_name']) else ""
                    if sku_number and model_name:
                        sku_dict[sku_number] = model_name
            else:
                # 기존 방식 (첫 번째, 두 번째 컬럼 사용)
                for _, row in sku_data.iterrows():
                    if pd.notna(row.iloc[0]):
                        sku_dict[str(row.iloc[0]).strip()] = str(row.iloc[1]).strip()
                    
            logger.info(f"SKU 딕셔너리 생성 완료: {len(sku_dict)} 항목")
            return sku_dict
            
        except Exception as e:
            logger.error(f"SKU 딕셔너리 생성 중 오류: {str(e)}")
            return {}
    
    @staticmethod
    def calculate_sum_formula(ws, target_col: str, source_cols: List[str], start_row: int = 2):
        """
        합계 수식 계산
        
        Args:
            ws: 워크시트
            target_col: 대상 열
            source_cols: 원본 열 리스트
            start_row: 시작 행 번호
        """
        try:
            last_row = ws.max_row
            processed_count = 0
            error_count = 0
            
            logger.info(f"합계 계산 시작: {target_col} = {' + '.join(source_cols)}")
            
            for row_num in range(start_row, last_row + 1):
                sum_value = 0
                valid_values = 0
                debug_info = []
                
                for col in source_cols:
                    cell_value = ws[f'{col}{row_num}'].value
                    debug_info.append(f"{col}{row_num}={cell_value}({type(cell_value)})")
                    
                    if cell_value is not None:
                        # DataTypeHandler를 사용하여 정수 변환
                        numeric_value = DataTypeHandler.to_integer(cell_value)
                        if numeric_value is not None:
                            sum_value += numeric_value
                            valid_values += 1
                            logger.debug(f"행 {row_num} {col}열: {cell_value} -> {numeric_value} (추가됨)")
                        else:
                            logger.debug(f"행 {row_num} {col}열: {cell_value} -> 변환 실패 (숫자가 아님)")
                
                # 유효한 값이 하나라도 있으면 결과 저장
                if valid_values > 0:
                    ws[f'{target_col}{row_num}'].value = sum_value
                    processed_count += 1
                    logger.debug(f"행 {row_num} 결과: {sum_value} (유효값 {valid_values}개)")
                else:
                    # 모든 값이 숫자가 아닌 경우 0으로 설정
                    ws[f'{target_col}{row_num}'].value = 0
                    error_count += 1
                    logger.debug(f"행 {row_num} 결과: 0 (모든 값이 숫자가 아님)")
                
                # 처음 몇 행만 상세 로깅
                if row_num <= start_row + 5:
                    logger.info(f"행 {row_num} 디버그: {' | '.join(debug_info)} -> 결과: {ws[f'{target_col}{row_num}'].value}")
                
            logger.info(f"합계 계산 완료: {target_col} - 처리된 행: {processed_count}, 오류 행: {error_count}")
            
        except Exception as e:
            logger.error(f"합계 계산 중 오류: {str(e)}")
            raise
    
    @staticmethod
    def transform_column_a_data(ws):
        """
        A열 데이터 변환
        G(beigebagel) -> [베이지베이글]G마켓-스마일
        A(beigebagel) -> [베이지베이글]옥션-스마일
        G(okokmart) -> [스마일배송]G마켓-스마일
        A(okokmart) -> [스마일배송][오케이마트]옥션-스마일
        G(clobuff1) -> [클로버프]G마켓-스마일
        A(clobuff1) -> [클로버프]옥션-스마일
        
        Args:
            ws: 워크시트
        """
        try:
            last_row = ws.max_row
            transformed_count = 0
            
            # 매핑 규칙 정의
            mapping_rules = {
                'beigebagel': {
                    'G': '[베이지베이글]G마켓-스마일',
                    'A': '[베이지베이글]옥션-스마일'
                },
                'okokmart': {
                    'G': '[스마일배송]G마켓-스마일',
                    'A': '[스마일배송][오케이마트]옥션-스마일'
                },
                'clobuff1': {
                    'G': '[클로버프]G마켓-스마일',
                    'A': '[클로버프]옥션-스마일'
                }
            }
            
            logger.info("A열 데이터 변환 시작")
            
            for row_num in range(2, last_row + 1):  # 헤더 제외하고 2행부터
                cell_value = ws[f'A{row_num}'].value
                
                if cell_value and isinstance(cell_value, str):
                    # 패턴 매칭: G(beigebagel) 또는 A(okokmart) 형태
                    import re
                    pattern = r'^([AG])\(([^)]+)\)$'
                    match = re.match(pattern, cell_value.strip())
                    
                    if match:
                        site_code = match.group(1)  # G 또는 A
                        store_code = match.group(2)  # beigebagel, okokmart, clobuff1 등
                        
                        # 매핑 규칙에서 찾기
                        if store_code in mapping_rules and site_code in mapping_rules[store_code]:
                            new_value = mapping_rules[store_code][site_code]
                            ws[f'A{row_num}'].value = new_value
                            transformed_count += 1
                            logger.debug(f"행 {row_num}: {cell_value} -> {new_value}")
                        else:
                            logger.debug(f"행 {row_num}: 매핑 규칙 없음 - {cell_value}")
                    else:
                        logger.debug(f"행 {row_num}: 패턴 불일치 - {cell_value}")
            
            logger.info(f"A열 데이터 변환 완료: {transformed_count}개 변환됨")
            
        except Exception as e:
            logger.error(f"A열 데이터 변환 중 오류: {str(e)}")
            raise
    
    @staticmethod
    def update_smile_macro_headers(ws):
        """
        스마일배송 매크로 헤더 업데이트
        
        Args:
            ws: 워크시트
        """
        try:
            logger.info("스마일배송 매크로 헤더 업데이트 시작")
            
            # 헤더 매핑 정보
            header_mapping = {
                5: "금액[배송비미포함]",   # E열
                12: "제품명"            # L열
            }
            
            # 헤더 업데이트
            for col_num, header_value in header_mapping.items():
                ws.cell(row=1, column=col_num, value=header_value)
                logger.debug(f"컬럼 {col_num} 헤더 업데이트: {header_value}")
            
            logger.info("스마일배송 매크로 헤더 업데이트 완료")
            
        except Exception as e:
            logger.error(f"스마일배송 매크로 헤더 업데이트 중 오류: {str(e)}")
            raise
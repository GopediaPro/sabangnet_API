import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Tuple, Any
from utils.logs.sabangnet_logger import get_logger
from utils.excels.excel_handler import ExcelHandler
from utils.macros.smile.smile_common_utils import SmileCommonUtils
from repository.smile_erp_data_repository import SmileErpDataRepository
from repository.smile_settlement_data_repository import SmileSettlementDataRepository

logger = get_logger(__name__)


class SmileMacroHandler:
    """
    스마일배송 매크로 핸들러
    VBA 매크로를 Python으로 변환한 처리 로직
    """
    
    def __init__(self, ws, wb=None, erp_repository: Optional[SmileErpDataRepository] = None, 
                 settlement_repository: Optional[SmileSettlementDataRepository] = None):
        self.ws = ws
        self.wb = wb
        self.erp_repository = erp_repository
        self.settlement_repository = settlement_repository
        self.logger = logger
    
    @classmethod
    def from_file(cls, file_path, sheet_index=0, erp_repository=None, settlement_repository=None):
        """
        파일에서 매크로 핸들러 생성
        
        Args:
            file_path: Excel 파일 경로
            sheet_index: 시트 인덱스 (기본값: 0)
            erp_repository: ERP 데이터 리포지토리
            settlement_repository: 정산 데이터 리포지토리
            
        Returns:
            SmileMacroHandler: 매크로 핸들러 인스턴스
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.worksheets[sheet_index]
        return cls(ws, wb, erp_repository, settlement_repository)
    
    def save_file(self, file_path):
        """
        파일 저장
        
        Args:
            file_path: 저장할 파일 경로
        """
        if file_path.endswith('_매크로_완료.xlsx'):
            output_path = file_path
        else:
            output_path = file_path.replace('.xlsx', '_매크로_완료.xlsx')
        self.wb.save(output_path)
        return output_path
    
    async def process_stage_1_to_5(self, erp_data: pd.DataFrame, settlement_data: pd.DataFrame):
        """
        1-5단계 처리
        
        Args:
            erp_data: ERP 데이터 DataFrame
            settlement_data: 정산 데이터 DataFrame
            
        Returns:
            bool: 처리 성공 여부
        """
        try:
            self._stage_1_basic_formatting()
            await self._stage_2_erp_matching(erp_data, settlement_data)
            self._stage_3_delete_colored_rows()
            self._stage_4_delete_conditioned_rows()
            self._stage_5_final_processing()
            return True
        except Exception as e:
            self.logger.error(f"1-5단계 처리 중 오류: {str(e)}")
            return False
    
    def get_an_column_value(self) -> Optional[str]:
        """
        AN 칼럼의 값을 가져와서 출력
        
        Returns:
            Optional[str]: AN 칼럼의 값, 오류 시 None
        """
        try:
            an_cell_value = self.ws['AN'].value
            if isinstance(an_cell_value, tuple):
                an_cell_value = str(an_cell_value)
            self.logger.info(f"macro_handler AN 칼럼 값: {an_cell_value}")
            return an_cell_value
        except Exception as e:
            self.logger.warning(f"AN 칼럼 값 출력 중 오류: {str(e)}")
            return None

    def print_all_columns_and_values(self, max_rows: int = 10, max_cols: int = 50) -> Dict[str, Any]:
        """
        전체 칼럼과 값을 출력하여 중간 점검
        
        Args:
            max_rows: 출력할 최대 행 수 (기본값: 10)
            max_cols: 출력할 최대 열 수 (기본값: 50)
            
        Returns:
            Dict[str, Any]: 칼럼 정보와 데이터
        """
        try:
            self.logger.info("=== 전체 칼럼 및 값 중간 점검 시작 ===")
            
            # 워크시트 정보
            total_rows = self.ws.max_row
            total_cols = self.ws.max_column
            self.logger.info(f"전체 행 수: {total_rows}, 전체 열 수: {total_cols}")
            
            # 헤더 정보 수집
            headers = {}
            for col in range(1, min(total_cols + 1, max_cols + 1)):
                col_letter = get_column_letter(col)
                header_value = self.ws[f'{col_letter}1'].value
                headers[col_letter] = header_value
                self.logger.info(f"열 {col_letter}: {header_value}")
            
            # 데이터 수집 (헤더 제외, 처음 몇 행만)
            data_rows = []
            for row in range(2, min(total_rows + 1, max_rows + 2)):
                row_data = {}
                for col in range(1, min(total_cols + 1, max_cols + 1)):
                    col_letter = get_column_letter(col)
                    cell_value = self.ws[f'{col_letter}{row}'].value
                    row_data[col_letter] = cell_value
                
                data_rows.append(row_data)
                
                # 각 행의 데이터 출력
                self.logger.info(f"=== 행 {row} 데이터 ===")
                for col_letter, value in row_data.items():
                    if value is not None:  # None이 아닌 값만 출력
                        self.logger.info(f"  {col_letter}: {value}")
            
            # 특별히 AN 칼럼 값도 별도로 출력
            try:
                an_cell = self.ws['AN']
                if isinstance(an_cell, tuple):
                    an_cell_value = str(an_cell)
                else:
                    an_cell_value = an_cell.value
                self.logger.info(f"=== AN 칼럼 값: {an_cell_value} ===")
            except Exception as e:
                self.logger.warning(f"AN 칼럼 값 출력 중 오류: {str(e)}")
            
            self.logger.info("=== 전체 칼럼 및 값 중간 점검 완료 ===")
            
            return {
                "total_rows": total_rows,
                "total_cols": total_cols,
                "headers": headers,
                "data_rows": data_rows,
                "an_column_value": an_cell_value if 'an_cell_value' in locals() else None
            }
            
        except Exception as e:
            self.logger.error(f"전체 칼럼 및 값 출력 중 오류: {str(e)}")
            return {}

    def process_stage_6_to_8(self, sku_data: pd.DataFrame):
        """
        6-8단계 처리
        
        Args:
            sku_data: SKU 데이터 DataFrame
            
        Returns:
            bool: 처리 성공 여부
        """
        try:
            self._stage_6_column_copy_and_format()
            self._stage_7_sku_processing(sku_data)
            # 전체 칼럼 및 값 중간 점검
            self.print_all_columns_and_values()
            self._stage_8_filter_and_sort()
            return True
        except Exception as e:
            self.logger.error(f"6-8단계 처리 중 오류: {str(e)}")
            return False
    
    def _stage_1_basic_formatting(self):
        """1단계: 기본 서식 처리"""
        # 공통 유틸리티 사용
        SmileCommonUtils.apply_basic_formatting(self.ws, font_size=9, row_height=15)
        SmileCommonUtils.delete_colored_rows(self.ws)
        
        # H열 자동 맞춤
        self.ws.column_dimensions['H'].auto_size = True
    
    async def _stage_2_erp_matching(self, erp_data: pd.DataFrame, settlement_data: pd.DataFrame):
        """2단계: I, J열 삽입 및 ERP 매칭"""
        # I, J열 삽입
        SmileCommonUtils.insert_columns(self.ws, 9, 2, ["ERP매칭", "정산여부"])
        
        # ERP 매칭 및 정산 여부 확인
        x_count = 0
        erp_count = 0
        settlement_count = 0
        
        for row_num in range(2, self.ws.max_row + 1):
            user_id = self.ws[f'A{row_num}'].value
            order_num = self.ws[f'H{row_num}'].value
            
            if order_num and user_id:
                # 데이터베이스에서 ERP 값과 정산 여부 조회
                erp_value = await self._find_erp_value(order_num)
                settlement_value = await self._check_settlement(order_num)
                # self.logger.info(f"행 {row_num} (사용자ID: {user_id}): ERP='{erp_value}', 정산='{settlement_value}', 주문번호='{order_num}'")
            else:
                erp_value = ""
                settlement_value = "X"
            
            # ERP 매칭 결과 설정
            self.ws[f'I{row_num}'].value = erp_value
            self.ws[f'J{row_num}'].value = settlement_value
            
            # 디버깅: ERP 값 확인 (처음 10행만)
            if row_num <= 10:
                self.logger.info(f"행 {row_num} ERP 값: '{erp_value}'")
            
            # 카운트
            if erp_value == "X":
                x_count += 1
            elif erp_value:
                erp_count += 1
            
            if settlement_value == "정산":
                settlement_count += 1
        
        self.logger.info(f"ERP 매칭 결과 - X: {x_count}건, ERP: {erp_count}건, 정산: {settlement_count}건")
    
    def _stage_3_delete_colored_rows(self):
        """3단계: 색이 있는 행 삭제"""
        # 공통 유틸리티를 사용하여 색이 있는 행 삭제
        SmileCommonUtils.delete_colored_rows(self.ws)
        self.logger.info("3단계: 색이 있는 행 삭제 완료")
    
    def _stage_4_delete_conditioned_rows(self):
        """4단계: 특정 조건 행 삭제"""
        rows_to_delete = []
        for row_num in range(2, self.ws.max_row + 1):
            erp_match = self.ws[f'I{row_num}'].value
            settlement = self.ws[f'J{row_num}'].value
            
            # 더 구체적인 조건으로 수정
            # ERP 매칭이 되고 정산이 완료된 경우만 삭제
            if erp_match == "ERP" and settlement == "정산":
                rows_to_delete.append(row_num)
                # self.logger.info(f"행 {row_num}: ERP=ERP, 정산=정산 (삭제 대상)")
            # ERP 매칭이 안되고 정산도 안된 경우는 보존 (데이터 검증 필요)
            elif erp_match == "X" and settlement == "X":
                self.logger.info(f"행 {row_num}: ERP=X, 정산=X (검증 필요 - 보존)")
        
        # 삭제될 행 수 로깅
        self.logger.info(f"4단계에서 삭제될 행 수: {len(rows_to_delete)} / 총 {self.ws.max_row - 1} 행")
        
        # 역순으로 삭제
        for row_num in reversed(rows_to_delete):
            self.ws.delete_rows(row_num)
        
        self.logger.info(f"4단계 처리 후 남은 행 수: {self.ws.max_row - 1}")
    
    def _stage_5_final_processing(self):
        """5단계: 최종 처리"""
        # I, J열 삭제 (ERP매칭, 정산여부 열 삭제)
        SmileCommonUtils.delete_columns(self.ws, 9, 2)
        
        # D열 오른쪽에 새 E열 삽입
        SmileCommonUtils.insert_columns(self.ws, 5, 1, ['금액[배송비미포함]'])
        
        # E1 셀에 노란색 배경
        SmileCommonUtils.set_cell_background(self.ws, "E1", "FFFF00")
        
        # B열(정산예정금) + C열(서비스 이용료) 더한 값 계산하여 E열에 입력
        # 정산예정금과 서비스 이용료의 합계를 계산
        # self.logger.info("5단계: B열(정산예정금) + C열(서비스 이용료) 합계 계산 시작")
        
        # 디버깅: 처음 몇 행의 B, C열 값 확인
        # for row_num in range(2, min(7, self.ws.max_row + 1)):
        #     b_value = self.ws[f'B{row_num}'].value
        #     c_value = self.ws[f'C{row_num}'].value
        #     self.logger.info(f"행 {row_num}: B열={b_value}({type(b_value)}), C열={c_value}({type(c_value)})")
        
        SmileCommonUtils.calculate_sum_formula(self.ws, "E", ["B", "C"], 2)
        self.logger.info("5단계: 합계 계산 완료")
    
    def _stage_7_sku_processing(self, sku_data: pd.DataFrame):
        """7단계: SKU 분해 및 모델명 조합"""
        # AA 뒤에 AB, AC, AD, AE 추가
        SmileCommonUtils.insert_columns(self.ws, 28, 4, ['SKU1번호', 'SKU1수량', 'SKU2번호', 'SKU2수량'])
        
        # 헤더 확인 (디버깅용)
        self.logger.info("=== 열 삽입 후 헤더 확인 ===")
        SmileCommonUtils.print_worksheet_headers(self.ws, max_cols=70, show_empty=True)
        
        # self.logger.info(f"sku_data: {sku_data}")
        
        # SKU 데이터를 딕셔너리로 변환
        sku_dict = SmileCommonUtils.create_sku_dictionary(sku_data)
        # self.logger.info(f"sku_dict: {sku_dict}")
        
        # AA 분해 및 모델명 조합
        for row_num in range(2, self.ws.max_row + 1):
            af_value = self.ws[f'AF{row_num}'].value
            if not af_value:
                continue
            
            result_text = ""
            af_str = str(af_value).strip()
            self.logger.info(f"AF 값: {af_str}")
            
            # SKU 패턴 분리 (예: "865797/1개 865798/1개", "832168/3개" 등)
            sku_combinations = []
            
            # 공백으로 분리하여 여러 SKU 조합 처리
            if " " in af_str:
                parts = af_str.split(" ")
                for part in parts:
                    if "/" in part:
                        sku_combinations.append(part.strip())
            else:
                # 단일 SKU인 경우
                if "/" in af_str:
                    sku_combinations.append(af_str)
            
            # 각 SKU 조합을 처리
            processed_skus = []
            for i, sku_combo in enumerate(sku_combinations):
                if "/" in sku_combo:
                    sku_parts = sku_combo.split("/")
                    if len(sku_parts) == 2:
                        sku_number = sku_parts[0].strip()
                        quantity = sku_parts[1].strip()
                        
                        # 모델명 찾기
                        if sku_number in sku_dict:
                            model_name = sku_dict[sku_number]
                            # 수량이 "1개"가 아닌 경우 수량 추가
                            if quantity != "1개":
                                processed_sku = f"{model_name} {quantity}"
                            else:
                                processed_sku = model_name
                            processed_skus.append(processed_sku)
                        else:
                            # SKU를 찾을 수 없는 경우 원본 값 사용
                            processed_skus.append(sku_combo)
                            self.logger.warning(f"SKU 번호 '{sku_number}'를 찾을 수 없습니다.")
            
            # 결과 조합
            if processed_skus:
                result_text = " + ".join(processed_skus)
            
            # AA, AB, AC, AD 컬럼에 분리된 값 저장
            if len(sku_combinations) >= 1:
                sku1_parts = sku_combinations[0].split("/") if "/" in sku_combinations[0] else ["", ""]
                self.ws[f'AB{row_num}'].value = sku1_parts[0].strip() if len(sku1_parts) > 0 else ""
                self.ws[f'AC{row_num}'].value = sku1_parts[1].strip() if len(sku1_parts) > 1 else ""
            
            if len(sku_combinations) >= 2:
                sku2_parts = sku_combinations[1].split("/") if "/" in sku_combinations[1] else ["", ""]
                self.ws[f'AD{row_num}'].value = sku2_parts[0].strip() if len(sku2_parts) > 0 else ""
                self.ws[f'AE{row_num}'].value = sku2_parts[1].strip() if len(sku2_parts) > 1 else ""
            
            # AE 컬럼에 최종 결과 저장
            self.ws[f'L{row_num}'].value = result_text
    
    def _stage_6_column_copy_and_format(self):
        """6단계: 열 복사 및 서식"""
        # K열 뒤에 L열 삽입
        SmileCommonUtils.insert_columns(self.ws, 12, 1, ['제품명'])
        
        # AF열 → L열로 복사
        # SmileCommonUtils.copy_column_values(self.ws, "AF", "L")
        
        # L1 셀 노란색 배경
        SmileCommonUtils.set_cell_background(self.ws, "L1", "FFFF00")
    
    def _stage_8_filter_and_sort(self):
        """8단계: 필터 및 정렬"""
        # 전체 필터 적용
        SmileCommonUtils.apply_auto_filter(self.ws)
        
        # 정렬: A → S (A기준 정렬)
        # 데이터를 DataFrame으로 변환하여 정렬
        data = []
        headers = []
        
        for row in self.ws.iter_rows(values_only=True):
            if not headers:
                headers = row
            else:
                data.append(row)
        
        df = pd.DataFrame(data, columns=headers)
        df_sorted = SmileCommonUtils.sort_dataframe(df, [headers[0], headers[50]], [True, True])  # A, S 컬럼
        
        # 정렬된 데이터를 다시 워크시트에 쓰기
        for row_idx, row_data in enumerate(df_sorted.values, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                self.ws.cell(row=row_idx, column=col_idx, value=value)
    
    async def _find_erp_value(self, order_no: str) -> str:
        """데이터베이스에서 ERP 값 찾기"""
        try:
            if not self.erp_repository:
                self.logger.warning("ERP 리포지토리가 설정되지 않았습니다. 빈 문자열을 반환합니다.")
                return ""
            
            # Excel에서 읽은 값이 숫자일 수 있으므로 문자열로 변환
            order_no_str = str(order_no) if order_no is not None else ""
            
            self.logger.debug(f"ERP 조회: 주문번호='{order_no_str}' (원본: '{order_no}')")
            
            erp_data = await self.erp_repository.get_erp_data_by_order_number(order_no_str)
            if erp_data and erp_data.erp_code:
                self.logger.debug(f"ERP 데이터 찾음: 주문번호='{order_no_str}' -> ERP코드='{erp_data.erp_code}'")
                return erp_data.erp_code
            else:
                self.logger.debug(f"ERP 데이터 없음: 주문번호='{order_no_str}'")
                return "X"
        except Exception as e:
            self.logger.error(f"ERP 값 조회 중 오류: {str(e)}")
            return ""
    
    async def _check_settlement(self, order_no: str) -> str:
        """데이터베이스에서 정산 여부 확인"""
        try:
            if not self.settlement_repository:
                self.logger.warning("정산 리포지토리가 설정되지 않았습니다. 'X'를 반환합니다.")
                return "X"
            
            # Excel에서 읽은 값이 숫자일 수 있으므로 문자열로 변환
            order_no_str = str(order_no) if order_no is not None else ""
            
            settlement_data_list = await self.settlement_repository.get_settlement_data_by_order_number(order_no_str)
            if settlement_data_list:
                return "정산"
            return "X"
        except Exception as e:
            self.logger.error(f"정산 여부 확인 중 오류: {str(e)}")
            return "X" 
from openpyxl.styles import Font, Alignment
from utils.excels.excel_handler import ExcelHandler
from utils.excels.excel_column_handler import ExcelColumnHandler


class ERPEtcSiteMacro:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.ex = ExcelHandler.from_file(file_path)
        self.ws = self.ex.ws
        self.wb = self.ex.wb
        self.order_set = set()
        self.toss_order_info = {}

    def etc_site_macro_run(self):
        col_h = ExcelColumnHandler()
        for row in range(2, self.ws.max_row + 1):
            self._overlap_by_site_column(self.ws[f"B{row}"], row)
            self._toss_process_column(self.ws[f"B{row}"], row)
            self._order_num_by_site_column(self.ws[f"B{row}"])
            self._kakao_jeju_process_column(
                self.ws[f"B{row}"], self.ws[f"J{row}"], self.ws[f"F{row}"])
            col_h.h_i_column(self.ws[f"H{row}"])
            col_h.h_i_column(self.ws[f"I{row}"])
        
        self._toss_order_info_process()

        # 시트 설정
        sheets_name = ["OK", "IY", "BB"]
        site_to_sheet = {
            "오케이마트": "OK",
            "아이예스": "IY",
            "베이지베이글": "BB",
        }

        # 정렬 기준: 2번째 컬럼(B) → 3번째 컬럼(C) 순으로 정렬 2025-07-17 정렬순서 조정, 3번쨰 컬럼은 내림차순되게 수정
        sort_columns = [2, 3, -4, 5, 6]
        print("시트별 정렬, 시트 분리 시작...")
        headers, data = self.ex.preprocess_and_update_ws(self.ws, sort_columns)
        self.ex.split_and_write_ws_by_site(
            wb=self.wb,
            headers=headers,
            data=data,
            sheets_name=sheets_name,
            site_to_sheet=site_to_sheet,
            site_col_idx=2,
        )
        print("시트별 정렬, 시트 분리 완료")

        print("시트별 서식, 디자인 적용 시작...")
        for ws in self.wb.worksheets:
            self.ex.set_header_style(ws)
            if ws.max_row <= 1:
                continue
            for row in range(2, ws.max_row + 1):
                col_h.a_value_column(ws[f"A{row}"])
                col_h.d_column(ws[f"D{row}"],ws[f"U{row}"], ws[f"V{row}"])
                col_h.e_column(ws[f"E{row}"])
                col_h.f_column(ws[f"F{row}"])
                col_h.l_column(ws[f"L{row}"])
                col_h.convert_int_column(ws[f"P{row}"])
                self._v_column_red_font(ws[f"V{row}"])
            print(f"[{ws.title}] 서식 및 디자인 적용 완료")
        # 장바구니번호와 수취인주소 조합으로 그룹화 후 평균 금액 적용
        self._average_duplicate_cart_address_amounts()
        output_path = self.ex.save_file(self.file_path)
        print(f"✓ 기타 사이트 ERP 자동화 완료! 최종 파일: {output_path}")
        return output_path

    def _overlap_by_site_column(self, b_cell, row):
        """
        각 사이트별 중복 주문 처리
        args:
            b_cell: 주문 번호 셀
            row: 행 번호
        """
        b_cell_text = str(b_cell.value)
        if "오늘의집" in b_cell_text:
            self.ws[f"V{row}"].value = 0
        elif "톡스토어" in b_cell_text:
            order_key = str(self.ws[f"E{row}"].value).strip()
            if order_key:
                if order_key in self.order_set:
                    self.ws[f"V{row}"].value = 0
                else:
                    self.order_set.add(order_key)
        elif any(site in b_cell_text for site in ["롯데온", "보리보리", "스마트스토어"]):
            order_key = str(self.ws[f"E{row}"].value).strip()
            if order_key:
                if order_key in self.order_set:
                    self.ws[f"V{row}"].value = 0
                else:
                    self.order_set.add(order_key)

    def _toss_process_column(self, cell, row):
        """
        토스 주문 처리
        args:
            b_cell: 주문 번호 셀
            row: 행 번호
        """
        b_cell_text = str(cell.value)
        if "토스" in b_cell_text:
            order_key = str(self.ws[f"E{row}"].value).strip()
            if order_key:
                u_value = self.ws[f"U{row}"].value
                u_value = float(u_value) if u_value is not None else 0
                if order_key not in self.toss_order_info:
                    self.toss_order_info[order_key] = [0.0, row]
                self.toss_order_info[order_key][0] += u_value
                self.ws[f"V{row}"].value = 0

    def _toss_order_info_process(self):
        """
        토스 주문 30000원 이하 처리
        """
        for _, (total, first_row) in self.toss_order_info.items():
            if total <= 30000:
                self.ws[f"V{first_row}"].value = 3000

    def _order_num_by_site_column(self, cell):
        """
        사이트별 주문 번호 처리
        args:
            cell: 주문 번호 셀
        """
        site_list = ["에이블리", "오늘의집", "쿠팡", "텐바이텐",
                     "NS홈쇼핑", "그립", "보리보리", "카카오선물하기", "톡스토어", "토스"]
        cell_text = str(cell.value)
        if any(site in cell_text for site in site_list):
            order_num = cell.value
            if order_num is not None and str(order_num).replace('.', '').isdigit():
                cell.value = int(float(order_num))

    def _kakao_jeju_process_column(self, b_cell, j_cell, f_cell):
        """
        카카오 제주 주문 처리
        args:
            b_cell: 주문 번호 셀
            j_cell: 주소 셀
            f_cell: 주문 번호 셀
        """
        b_text = str(b_cell.value)
        j_text = str(j_cell.value)
        if "카카오" in b_text and "제주" in j_text:
            f_text = str(f_cell.value)
            if "[3000원 연락해야함]" not in f_text:
                f_cell.value = f_text + "[3000원 연락해야함]"
                f_cell.font = Font(color="FF0000", bold=True)
                

    def _v_column_red_font(self, v_cell):
        """
        V 컬럼 빨간색 글씨 처리
        args:
            v_cell: V 컬럼 셀
        """
        if v_cell.value == 0:
            v_cell.font = Font(color="FF0000", bold=True)
            v_cell.alignment = Alignment(horizontal='right')

    def _average_duplicate_cart_address_amounts(self):
        """
        Q셀(장바구니번호)과 J셀(수취인주소)이 같은 값들의 D셀(금액)을 평균내는 메서드
        """
        # 장바구니번호와 수취인주소 조합으로 그룹화
        duplicate_groups = {}
        
        for row in range(2, self.ws.max_row + 1):
            cart_number = str(self.ws[f"Q{row}"].value).strip()
            address = str(self.ws[f"J{row}"].value).strip()
            amount = self.ws[f"D{row}"].value
            
            # 빈 값이 아닌 경우에만 처리
            if cart_number and cart_number != "None" and address and address != "None":
                key = f"{cart_number}_{address}"
                
                if key not in duplicate_groups:
                    duplicate_groups[key] = []
                
                duplicate_groups[key].append({
                    'row': row,
                    'amount': int(amount) if amount is not None else 0
                })
        
        # 2개 이상인 그룹에 대해 평균 계산 및 적용
        for key, group in duplicate_groups.items():
            if len(group) >= 2:
                total_amount = sum(item['amount'] for item in group)
                average_amount = total_amount / len(group)
                
                # 모든 행에 평균값 적용
                for item in group:
                    self.ws[f"D{item['row']}"].value = average_amount
                
                print(f"장바구니번호-주소 조합 '{key}'의 {len(group)}개 행에 평균 금액 {average_amount:.2f} 적용")

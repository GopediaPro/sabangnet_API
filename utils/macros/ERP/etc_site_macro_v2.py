from openpyxl.cell import Cell
from openpyxl.styles import Font, Alignment
from utils.excels.excel_handler import ExcelHandler
from utils.excels.excel_column_handler import ExcelColumnHandler
from utils.macros.ERP.utils import average_duplicate_order_address_amounts
from utils.logs.sabangnet_logger import get_logger

logger = get_logger(__name__)

class ERPEtcSiteMacroV2:
    def __init__(self, file_path: str, is_star: bool = False):
        self.file_path = file_path
        self.is_star = is_star
        self.ex = ExcelHandler.from_file(file_path)
        self.ws = self.ex.ws
        self.wb = self.ex.wb
        self.order_set = set()
        self.toss_order_info = {}

    def etc_site_macro_run(self):
        """
        개선된 프로세스 순서:
        1. 시트 생성
        2. 데이터 처리 (컬럼 처리, 토스 주문 처리 등)
        3. 스타배송 모드에서 평균 금액 적용
        4. 시트별로 데이터 분리
        5. 시트별 디자인 적용
        """
        logger.info("=== 기타 사이트 ERP 자동화 V2 시작 ===")
        
        # 1단계: 시트 설정 및 생성
        sheets_name = ["OK", "IY", "BB"]
        site_to_sheet = {
            "오케이마트": "OK",
            "아이예스": "IY",
            "베이지베이글": "BB",
        }
        
        # 필요한 시트들이 없으면 생성
        self._ensure_sheets_exist(sheets_name)
        logger.info("✓ 시트 생성 완료")

        # 2단계: 데이터 처리
        logger.info("데이터 처리 시작...")
        col_h = ExcelColumnHandler()
        
        # 기본 데이터 처리
        for row in range(2, self.ws.max_row + 1):
            self._overlap_by_site_column(self.ws[f"B{row}"], row)
            self._overlap_kakao(row)
            self._toss_process_column(self.ws[f"B{row}"], row)
            self._order_num_by_site_column(self.ws[f"B{row}"])
            self.ws[f"H{row}"].value = self.ex.format_phone_number(self.ws[f"H{row}"].value)
            self.ws[f"I{row}"].value = self.ex.format_phone_number(self.ws[f"I{row}"].value)
        
        # 토스 주문 정보 처리
        self._toss_order_info_process()
        
        # D, U, V 컬럼 처리 (토스 주문 정보 처리 후)
        for row in range(2, self.ws.max_row + 1):
            col_h.d_column(self.ws[f"D{row}"], self.ws[f"U{row}"], self.ws[f"V{row}"])
        logger.info("✓ 기본 데이터 처리 완료")

        # 3단계: 스타배송 모드에서 평균 금액 적용
        if self.is_star:
            logger.info("스타배송 모드: 평균 금액 적용 중...")
            average_duplicate_order_address_amounts(self.ws)
            logger.info("✓ 평균 금액 적용 완료")

        # 4단계: 시트별로 데이터 분리
        logger.info("시트별 데이터 분리 시작...")
        sort_columns = [2, 3, -4, 5, 6]  # 정렬 기준
        headers, data = self.ex.preprocess_and_update_ws(self.ws, sort_columns)
        
        self.ex.split_and_write_ws_by_site(
            wb=self.wb,
            headers=headers,
            data=data,
            sheets_name=sheets_name,
            site_to_sheet=site_to_sheet,
            site_col_idx=2,
        )
        logger.info("✓ 시트별 데이터 분리 완료")

        # 5단계: 시트별 디자인 적용
        logger.info("시트별 서식, 디자인 적용 시작...")
        for ws in self.wb.worksheets:
            if ws.title == "Sheet":  # 기본 시트는 건너뛰기
                continue
                
            self.ex.set_header_style(ws)
            if ws.max_row <= 1:
                continue
                
            for row in range(2, ws.max_row + 1):
                col_h.a_value_column(ws[f"A{row}"])
                col_h.e_column(ws[f"E{row}"])
                col_h.f_column(ws[f"F{row}"])
                col_h.l_column(ws[f"L{row}"])
                col_h.convert_int_column(ws[f"P{row}"])
                self._v_column_red_font(ws[f"V{row}"])
            
            logger.info(f"✓ [{ws.title}] 서식 및 디자인 적용 완료")
        
        # 최종 파일 저장
        output_path = self.ex.save_file(self.file_path)
        logger.info(f"✓ 기타 사이트 ERP 자동화 V2 완료! 최종 파일: {output_path}")
        return output_path

    def _ensure_sheets_exist(self, sheets_name):
        """
        필요한 시트들이 존재하는지 확인하고 없으면 생성
        """
        existing_sheets = [ws.title for ws in self.wb.worksheets]
        
        for sheet_name in sheets_name:
            if sheet_name not in existing_sheets:
                self.wb.create_sheet(title=sheet_name)
                logger.info(f"  - {sheet_name} 시트 생성됨")

    def _overlap_by_site_column(self, b_cell, row):
        """
        각 사이트별 중복 주문 처리
        args:
            b_cell: 주문 번호 셀
            row: 행 번호
        """
        b_cell_text = str(b_cell.value)
        if "오늘의집" in b_cell_text:
            self.ws[f"V{row}"].value = 0
        elif "톡스토어" in b_cell_text:
            order_key = str(self.ws[f"E{row}"].value).strip()
            if order_key:
                if order_key in self.order_set:
                    self.ws[f"V{row}"].value = 0
                else:
                    self.order_set.add(order_key)
        elif any(site in b_cell_text for site in ["롯데온", "보리보리", "스마트스토어"]):
            order_key = str(self.ws[f"E{row}"].value).strip()
            if order_key:
                if order_key in self.order_set:
                    self.ws[f"V{row}"].value = 0
                else:
                    self.order_set.add(order_key)
    
    # 카카오톡스토어 배송비 중복 처리 (_overlap_by_site_column 로직 보완)
    def _overlap_kakao(self, row: int):
        """
        사이트 이름, 수취인명, 수취인주소, 우편번호 기준으로 중복값 확인 후 배송비 처리
        args:
            row: 행 번호
        """
        site_name = str(self.ws[f"B{row}"].value).strip()
        if "카카오톡스토어" in site_name:
            receiver_name = str(self.ws[f"C{row}"].value).strip()
            receiver_address = str(self.ws[f"J{row}"].value).strip()
            receiver_zip_code = str(self.ws[f"K{row}"].value).strip()
            
            # 키 생성
            order_key = f"{site_name}_{receiver_name}_{receiver_address}_{receiver_zip_code}"

            if not order_key:
                return

            if order_key in self.order_set:
                self.ws[f"V{row}"].value = 0
            else:
                self.order_set.add(order_key)

    def _toss_process_column(self, cell, row):
        """
        토스 주문 처리
        args:
            b_cell: 주문 번호 셀
            row: 행 번호
        """
        b_cell_text = str(cell.value)
        if "토스" in b_cell_text:
            order_key = str(self.ws[f"E{row}"].value).strip()
            if order_key:
                u_value = self.ws[f"U{row}"].value
                u_value = float(u_value) if u_value is not None else 0
                if order_key not in self.toss_order_info:
                    self.toss_order_info[order_key] = [0.0, row]
                self.toss_order_info[order_key][0] += u_value
                self.ws[f"V{row}"].value = 0

    def _toss_order_info_process(self):
        """
        토스 주문 30000원 이하 처리
        """
        for _, (total, first_row) in self.toss_order_info.items():
            if total <= 30000:
                self.ws[f"V{first_row}"].value = 3000

    def _order_num_by_site_column(self, cell):
        """
        사이트별 주문 번호 처리
        args:
            cell: 주문 번호 셀
        """
        site_list = ["에이블리", "오늘의집", "쿠팡", "텐바이텐",
                     "NS홈쇼핑", "그립", "보리보리", "카카오선물하기", "톡스토어", "토스"]
        cell_text = str(cell.value)
        if any(site in cell_text for site in site_list):
            order_num = cell.value
            if order_num is not None and str(order_num).replace('.', '').isdigit():
                cell.value = int(float(order_num))

    def _v_column_red_font(self, v_cell):
        """
        V 컬럼 빨간색 글씨 처리
        args:
            v_cell: V 컬럼 셀
        """
        if v_cell.value == 0:
            v_cell.font = Font(color="FF0000", bold=True)
            v_cell.alignment = Alignment(horizontal='right')

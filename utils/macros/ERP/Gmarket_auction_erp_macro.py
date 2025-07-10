import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from utils.excel_handler import ExcelHandler


class GmarketAuctionMacro:
    def __init__(self, file_path):
        self.ex = ExcelHandler.from_file(file_path)
        self.file_path = file_path
        self.ws = self.ex.ws
        self.wb = self.ex.wb
        self.last_row = self.ws.max_row
        self.last_column = self.ws.max_column
        self.right_alignment = Alignment(horizontal='right')
        self.center_alignment = Alignment(horizontal='center')
        self.light_blue_fill = PatternFill(
            start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        self.dark_green_fill = PatternFill(
            start_color="008000", end_color="008000", fill_type="solid")
        self.white_font = Font(name='맑은 고딕', size=9, color="FFFFFF", bold=True)
        self.headers = []
        self.df = None
        self.ws_map = {}

    def step_1_to_12(self):
        """
        [1~12단계] 자동화 실행
        """
        print("1~12단계 자동화 시작...")
        self._step_1()  # 정렬 먼저
        self._step_2()
        self._step_3()
        self._step_4()
        self._step_5()
        self._step_6()

        # 모든 시트에 서식 적용
        print("12단계: 모든 시트에 서식 적용 시작...")
        for ws in self.wb.worksheets:
            if ws.max_row <= 1:
                continue
            self._step_12(ws)

        output_path = self.ex.save_file(self.file_path)
        return output_path

    def _step_1(self):
        """
        [1단계] 데이터를 DataFrame으로 변환
        """
        # 데이터를 DataFrame으로 변환
        self.df = self.ex.to_dataframe(self.ws)
        self.headers = list(self.df.columns)

        # C열(인덱스 2), B열(인덱스 1) 순서로 정렬
        if len(self.df.columns) > 2:
            self.df = self.ex.sort_dataframe_by_c_b(
                self.df, c_col='수취인명', b_col='사이트')

            # 워크시트에 정렬된 데이터 덮어쓰기
            for row_idx, row_data in enumerate(self.df.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    self.ws.cell(row=row_idx, column=col_idx, value=value if value or value == 0 else "")

        print("1단계: C열, B열 순서 정렬 완료")


    def _step_2(self):
        """
        [2단계] 장바구니 중복 제거
        """

        basket_dict = {}

        # 첫 번째 패스: 바구니 번호별로 배송비가 있는 첫 번째 행 기록
        for row in range(2, self.last_row + 1):
            basket_no = str(self.ws[f'Q{row}'].value).strip(
            ) if self.ws[f'Q{row}'].value else ""
            shipping_cost = self.ws[f'V{row}'].value

            if basket_no and basket_no != "":
                if basket_no not in basket_dict:
                    # 배송비가 0이 아니고 비어있지 않은 경우만 기록
                    if shipping_cost and shipping_cost != 0:
                        basket_dict[basket_no] = row

        # 두 번째 패스: 중복 바구니의 배송비를 0으로 설정
        for row in range(2, self.last_row + 1):
            basket_no = str(self.ws[f'Q{row}'].value).strip(
            ) if self.ws[f'Q{row}'].value else ""

            if basket_no and basket_no != "":
                if basket_no in basket_dict:
                    # 첫 번째 발생 행이 아닌 경우 배송비를 0으로 설정
                    if basket_dict[basket_no] != row:
                        self.ws[f'V{row}'].value = 0

        print(f"2단계: {len(basket_dict)}개 바구니 중복 제거 완료")

    def _step_3(self):
        """
        [3단계] 색칠음영 제거
        """
        # 색칠 음영 제거 (A2:Z까지)
        self.ex.clear_fills_from_second_row()

        print("3단계: A열 순번 입력 및 배경색 제거 완료")


    def _step_4(self):
        """
        [4단계] 1행 가운데 정렬 + 진한 초록색 배경
        """
        headers = [self.ws.cell(row=1, column=col).value for col in range(
            1, self.ws.max_column + 1)]

        self.ex.set_header_style(
            self.ws, headers, self.dark_green_fill, self.white_font, self.center_alignment)
        print("4단계: 헤더 서식 완료")
    
    def _step_5(self):
        """ 
        [6단계] 시트 분리 준비
        """
         # 기존 시트 삭제
        sheets_name: list[str] = ["OK,CL,BB", "IY"]
        
        self.ws_map = self.ex.create_split_sheets(self.headers, sheets_name)
        
        # 지정한 워크시트의 헤더 행에 배경색, 폰트, 정렬을 일괄 적용
        self.ex.set_header_style(self.ws_map["OK,CL,BB"], self.headers,self.dark_green_fill, self.white_font, self.center_alignment)
        self.ex.set_header_style(self.ws_map["IY"], self.headers, self.dark_green_fill, self.white_font, self.center_alignment)
        
        print("5단계: 시트 분리 준비 완료")
            
    def _step_6(self):
        """ 
        [6단계] 데이터 분류 및 복사
        """
        ok_row = 2
        iy_row = 2
        font = Font(name='맑은 고딕', size=9)
        
        for row_data in self.df.itertuples(index=False):

            # B열(사이트 정보)에서 계정명 추출
            site_value = str(getattr(row_data, '사이트')) if pd.notna(getattr(row_data, '사이트')) else ""
            account_name = ""
            
            if "]" in site_value and site_value.startswith("["):
                try:
                    account_name = site_value[1:site_value.index("]")]
                except:
                    account_name = ""


            # 계정명에 따라 시트 분리
            if account_name in ["오케이마트", "클로버프", "베이지베이글"]:
                # OK,CL,BB 시트에 복사
                for col_idx, value in enumerate(row_data, 1):
                    cell = self.ws_map["OK,CL,BB"].cell(row=ok_row, column=col_idx, value=value)
                    cell.font = font
                self.ws_map["OK,CL,BB"].row_dimensions[ok_row].height = 15
                ok_row += 1
                
            elif account_name == "아이예스":
                # IY 시트에 복사
                for col_idx, value in enumerate(row_data, 1):
                    cell = self.ws_map["IY"].cell(row=iy_row, column=col_idx, value=value)
                    cell.font = font
                self.ws_map["IY"].row_dimensions[iy_row].height = 15
                iy_row += 1
        
        print(f"6단계: 데이터 분리 완료 (OK: {ok_row-2}행, IY: {iy_row-2}행)")
    
    def _step_7(self, ws):
        """
        [7단계] D열 수식 활성화 및 채우기
        """
        d2_formula = ws['D2'].value
        self.ex.autofill_d_column(ws=ws,
            start_row=2, end_row=ws.max_row, formula=d2_formula)
        print("7단계: D열 수식 처리 완료")

    def _step_8(self, ws=None):
        """
        [8단계] Q, P, M, W 숫자 변환 및 오른쪽 정렬
        """
        self.ex.set_column_alignment(ws)

        cols_names = ['Q', 'P', 'M', 'W']
        self.ex.convert_numeric_strings(
            ws=ws, start_row=2, end_row=self.last_row, cols=cols_names)

        print("8단계: Q, P, M, W 숫자 변환 및 오른쪽 정렬 완료")

    def _step_9(self, ws):
        """
        [9단계] F열 ' 1개' 제거
        """
        for row in range(2, ws.max_row + 1):
            ws[f'F{row}'].value = self.ex.clean_model_name(ws[f'F{row}'].value)

        print("9단계: F열 ' 1개' 제거 완료")

    def _step_10(self, ws):
        """
        [10단계] F열 모르겠는 셀 색칠음영 (하늘색)
        """
        self.ex.highlight_column(
            col='F', light_color=self.light_blue_fill, ws=ws, start_row=2, last_row=ws.max_roww)
        print("11단계: F열 조건부 하이라이트 완료")

    def _step_11(self, ws):
        """
        [11단계] L열 처리
        """
        red_font = Font(color="FF0000")

        for row in range(2, ws.max_row + 1):
            l_value = ws[f'L{row}'].value
            if l_value:
                l_value_str = str(l_value).strip()
                if l_value_str == "신용":
                    ws[f'L{row}'].value = ""
                elif l_value_str == "착불":
                    ws[f'L{row}'].font = red_font

        print("11단계: L열 처리 완료 (신용→공백, 착불→빨간색)")

    def _step_12(self, ws):
        """ 
        [12단계] 모든 시트에 서식 적용
        """

        # A열 수식 값 변환
        self.ex.set_row_number(ws)
        
        # 테두리 제거 & 격자 제거
        self.ex.clear_borders(ws)
        
        # D열 수식 활성화 및 채우기
        self._step_7(ws)
        
        # 기본 폰트 적용
        self.ex.set_basic_format(header_rgb="008000")
        
        # A, B, D, E, G열 정렬
        self.ex.set_column_alignment()
        
        # E열 숫자 변환 및 오른쪽 정렬
        for row in range(2, self.last_row + 1):
            e_value = ws[f'E{row}'].value
            if e_value and str(e_value).replace('.', '').replace('-', '').isdigit():
                ws[f'E{row}'].value = self.ex.convert_to_number(e_value)
            ws[f'E{row}'].alignment = self.right_alignment

        # Q, P, M, W 숫자 변환 및 오른쪽 정렬
        self._step_8(ws)

        # F열 '1개' 제거
        self._step_9(ws)

        # F열 모르겠는 셀 색칠음영 (하늘색)
        self._step_10(ws)

        # L열 처리
        self._step_11(ws)

        print(f"12단계: [{ws.title}] 서식 재적용 완료")
        
    

def gok_erp_automation_full(file_path):
    """
    G,옥 ERP 자동화 전체 프로세스 실행
    
    Args:
        file_path (str): 처리할 Excel 파일 경로
    """
    
    print("G,옥 ERP 자동화 전체 프로세스 시작...")
    
    # 1~12단계 실행
    class_instance = GmarketAuctionERP(file_path)
    final_file = class_instance.step_1_to_12()
    
    print(f"✓ G,옥 ERP 자동화 완료! 최종 파일: {final_file}")
    return final_file



from openpyxl.styles import Font, PatternFill, Alignment
import re
from utils.excel_handler import ExcelHandler


class BrandiMacro:
    def __init__(self, file_path):
        self.ex = ExcelHandler.from_file(file_path)
        self.file_path = file_path
        self.ws = self.ex.ws
        self.wb = self.ex.wb
        self.last_row = self.ws.max_row
        self.last_col = self.ws.max_column
        self.right_alignment = Alignment(horizontal='right')
        self.center_alignment = Alignment(horizontal='center')
        self.light_blue_fill = PatternFill(
            start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        self.dark_green_fill = PatternFill(
            start_color="008000", end_color="008000", fill_type="solid")
        self.white_font = Font(name='맑은 고딕', size=9, color="FFFFFF", bold=True)
        self.pattern = re.compile(r'(?<!\w)\d+개$')
        self.headers = []
        self.df = None

    def step_1_to_11(self) -> str:
        print("브랜디 ERP 자동화 1~11단계 처리 시작...")
        self._step_1() # 10단계 정렬 먼저 후 적용
        self._step_2()
        self._step_3()
        self._step_4()
        self._step_5()
        self._step_6()
        self._step_7()
        self._step_8()
        self._step_9()
        self._step_10()
        self._step_11()
         
        output_path = self.ex.save_file(self.file_path)
        print(f"브랜디 ERP 자동화 1~11단계 모든 처리 완료!\n처리된 파일: {output_path}")
        return output_path
    
    def _step_1(self):
        """
        1단계: C열 기준 오름차순 정렬
        """
        # 데이터를 DataFrame으로 변환
        self.df = self.ex.to_dataframe(self.ws)
        self.headers = list(self.df.columns)

        # 데이터 정렬
        if len(self.df.columns) > 2:
            self.df = self.df.sort_values(by=['수취인명']).reset_index(drop=True)

            # 워크시트에 정렬된 데이터 덮어쓰기
            for row_idx, row_data in enumerate(self.df.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    self.ws.cell(row=row_idx, column=col_idx, value=value if value or value == 0 else "")

        print("1단계: C열 기준 정렬 완료")

    def _step_2(self):
        """
        2단계: 서식 적용
        """
        self.ex.set_basic_format(header_rgb="006100")
        print(f"2단계: 서식 적용 완료 ")

    def _step_3(self):
        """
        3단계: D열 수식 적용 (AutoFill)
        """
        d2_formula = self.ws['D2'].value
        self.ex.autofill_d_column(
            start_row=2, end_row=self.last_row, formula=d2_formula)
        print("3단계: D열 수식 적용 완료")

    def _step_4(self):
        """
        4단계: F열 " 1개" 제거
        """
        for row in range(2, self.last_row + 1):
            self.ws[f'F{row}'].value = self.ex.clean_model_name(
                self.ws[f'F{row}'].value)
        print("4단계: F열 ' 1개' 제거 완료")

    def _step_5(self):
        """
        5단계: 색 채우기 제거
        """
        for row in range(2, self.last_row + 1):
            for col in range(1, self.last_col + 1):
                self.ws.cell(row=row, column=col).fill = PatternFill(
                    fill_type=None)
        print("5단계: 배경색 제거 완료")

    def _step_6(self):
        """
        6단계: F열 조건부 연한 파란색 칠하기
        """
        for row in range(2, self.last_row + 1):
            f_value = self.ws[f'F{row}'].value
            if f_value:
                cell_value = str(f_value).strip()
                if cell_value.isdigit() or self.pattern.search(cell_value):
                    self.ws[f'F{row}'].fill = self.light_blue_fill
        print("6단계: F열 조건부 파란색 칠하기 완료")

    def _step_7(self):
        """
        7단계: 테두리 제거 & 격자 제거 + 연락처 포맷
        """
        # 테두리 제거 & 격자 제거
        self.ex.clear_borders()

        # 연락처 포맷
        for row in range(2, self.last_row + 1):
            for col_letter in ['H', 'I']:
                value = self.ws[f'{col_letter}{row}'].value
                if value:
                    self.ws[f'{col_letter}{row}'] = self.ex.format_phone_number(value)

        print("7단계: 테두리 제거 및 전화번호 포맷팅 완료")

    def _step_8(self):
        """
        [8단계] A열 순번 수식 입력
        """
        self.ex.set_row_number(self.ws, start_row=2, end_row=self.last_row)
        print("8단계: A열 순번 입력 완료")

    def _step_9(self):  
        """
        9단계: 제주 주소 안내문 + 서식 반영
        """
        for row in range(2, self.last_row + 1):
            j_value = self.ws[f'J{row}'].value
            if j_value and "제주" in str(j_value):
                self.ex.process_jeju_address(
                    self.ws, row, f_col='F', j_col='J')

        print(f"9단계: 제주 주소 처리 완료")

    def _step_10(self):
        """
        10단계: 정렬 설정
        """
        self.ex.set_column_alignment(self.ws)
        print("10단계: 정렬 설정 완료")


    def _step_11(self):
        """
        11단계: 워크시트의 문자열 숫자를 숫자 타입(int/float)으로 변환합니다.
        """
        self.ex.convert_numeric_strings(start_row=2, end_row=self.last_row, cols=("E", "P", "W"))

        print("11단계: 문자열 숫자 변환 완료")



def brandi_erp_macro_1_to_10(file_path):
    """
    브랜디 ERP 자동화 1~11단계 전체 자동처리 함수

    Args:
        file_path (str): 처리할 Excel 파일 경로
    """
    brandi = BrandiMacro(file_path)
    output_path = brandi.step_1_to_11()


    print(f"✓ 브랜디 ERP 자동화 1~11단계 모든 처리 완료!")
    print(f"처리된 파일: {output_path}")

    return output_path

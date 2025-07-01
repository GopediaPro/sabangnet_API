import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from collections import defaultdict

def other_site_macro_1_to_14(file_path):
    """
    Excel 매크로의 Step1_14_전체작업통합 함수를 Python으로 변환
    
    Args:
        file_path (str): 처리할 Excel 파일 경로
    """
    print(f"파일 경올 : {file_path}")
    # Excel 파일 로드
    workbook = openpyxl.load_workbook(file_path)
    ws = workbook.active
    
    # 폰트 및 행 높이 설정
    font = Font(name='맑은 고딕', size=9)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
    
    # 행 높이 설정
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 15
    
    # 첫 번째 행 배경색 설정 (녹색)
    green_fill = PatternFill(start_color="006100", end_color="006100", fill_type="solid")
    for cell in ws[1]:
        cell.fill = green_fill
    
    # D열 수식 설정 (D = U + V)
    last_row = ws.max_row
    for i in range(2, last_row + 1):
        ws[f'D{i}'].value = f'=U{i}+V{i}'
    
    # 데이터를 DataFrame으로 변환하여 정렬
    data = []
    headers = [cell.value for cell in ws[1]]
    
    for row in range(2, last_row + 1):
        row_data = [ws[f'{chr(65 + col)}{row}'].value for col in range(len(headers))]
        data.append(row_data)
    
    df = pd.DataFrame(data, columns=headers)
    
    # B열(인덱스 1), C열(인덱스 2) 기준으로 정렬
    df = df.sort_values(by=[df.columns[2], df.columns[1]])  # C열, B열 순서로 정렬
    
    # A열에 순번 재설정
    df.iloc[:, 0] = range(1, len(df) + 1)
    
    # 딕셔너리 초기화
    order_dict = {}
    u_sum_dict = defaultdict(float)
    row_list_dict = defaultdict(list)
    row_dict = defaultdict(list)
    
    # 사이트별 처리 로직
    for idx, row in df.iterrows():
        site_text = str(row.iloc[1])  # B열
        
        # 오늘의집 처리
        if "오늘의집" in site_text:
            df.at[idx, df.columns[21]] = 0  # V열 (인덱스 21)
            # 빨간색, 굵게 표시는 나중에 Excel에서 적용
        
        # 톡스토어 처리
        elif "톡스토어" in site_text:
            order_key = str(row.iloc[9]).strip()  # J열
            if order_key and order_key != "":
                if order_key in order_dict:
                    df.at[idx, df.columns[21]] = 0  # V열
                else:
                    order_dict[order_key] = True
        
        # 롯데온, 보리보리, 스마트스토어 처리
        elif any(site in site_text for site in ["롯데온", "보리보리", "스마트스토어"]):
            order_key = str(row.iloc[4]).strip()  # E열
            if order_key and order_key != "":
                if order_key in order_dict:
                    df.at[idx, df.columns[21]] = 0  # V열
                else:
                    order_dict[order_key] = True
    
    # 토스 처리
    for idx, row in df.iterrows():
        if "토스" in str(row.iloc[1]):  # B열
            order_key = str(row.iloc[4]).strip()  # E열
            if order_key and order_key != "":
                u_value = float(row.iloc[20]) if pd.notna(row.iloc[20]) else 0  # U열
                u_sum_dict[order_key] += u_value
                row_list_dict[order_key].append(idx)
    
    # 토스 주문 처리 (30000원 기준)
    for order_key, rows in row_list_dict.items():
        if u_sum_dict[order_key] > 30000:
            for row_idx in rows:
                df.at[row_idx, df.columns[21]] = 0  # V열
        else:
            for i, row_idx in enumerate(rows):
                df.at[row_idx, df.columns[21]] = 3000 if i == 0 else 0  # 첫 번째만 3000
    
    # 전화번호 포맷팅 (H열, I열)
    for col_idx in [7, 8]:  # H열(7), I열(8)
        for idx in df.index:
            phone_val = str(df.iloc[idx, col_idx]).strip()
            if (len(phone_val) == 11 and 
                phone_val.startswith('010') and 
                phone_val.isdigit()):
                formatted_phone = f"{phone_val[:3]}-{phone_val[3:7]}-{phone_val[7:]}"
                df.iloc[idx, col_idx] = formatted_phone
    
    # 사이트별 주문번호 숫자 변환
    site_list = ["에이블리", "오늘의집", "쿠팡", "텐바이텐", "NS홈쇼핑", 
                 "그립", "보리보리", "카카오선물하기", "톡스토어", "토스"]
    
    for idx, row in df.iterrows():
        site_text = str(row.iloc[1])  # B열
        for site in site_list:
            if site in site_text:
                row_dict[site].append(idx)
                break
    
    # 주문번호 숫자 변환 (E열)
    for site, indices in row_dict.items():
        for idx in indices:
            order_num = df.iloc[idx, 4]  # E열
            if pd.notna(order_num) and str(order_num).replace('.', '').isdigit():
                df.iloc[idx, 4] = int(float(order_num))
    
    # F열 처리 (상품명)
    regex_exact = re.compile(r'^\d+개?$', re.IGNORECASE)
    regex_repeat = re.compile(r'(\d+개\s*){2,}', re.IGNORECASE)
    
    for idx in df.index:
        f_val = str(df.iloc[idx, 5])  # F열
        
        # " 1개" 제거
        if " 1개" in f_val:
            f_val = f_val.replace(" 1개", "")
            df.iloc[idx, 5] = f_val
        
        # 정규식 패턴 확인 (하이라이트 표시용)
        if regex_exact.match(f_val) or regex_repeat.search(f_val):
            # 파란색 하이라이트 표시는 나중에 Excel에서 적용
            pass
    
    # L열 처리 (배송비 관련)
    for idx in df.index:
        l_val = str(df.iloc[idx, 11]).strip()  # L열
        if l_val == "신용":
            df.iloc[idx, 11] = ""
        elif l_val == "착불":
            # 빨간색, 굵게 표시는 나중에 Excel에서 적용
            pass
    
    # 카카오 + 제주 처리
    for idx in df.index:
        if ("카카오" in str(df.iloc[idx, 1]) and  # B열
            "제주" in str(df.iloc[idx, 9])):      # J열
            
            f_val = str(df.iloc[idx, 5])  # F열
            if "[3000원 연락해야함]" not in f_val:
                df.iloc[idx, 5] = f_val + " [3000원 연락해야함]"
    
    # DataFrame을 Excel로 다시 저장
    # 기존 시트 클리어
    ws.delete_rows(2, ws.max_row)
    
    # 새 데이터 추가
    for r_idx, (_, row) in enumerate(df.iterrows(), 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # 시트 분리 (14단계)
    site_filters = ["오케이마트", "아이예스", "베이지베이글"]
    site_codes = ["OK", "IY", "BB"]
    
    for i, (site_filter, site_code) in enumerate(zip(site_filters, site_codes)):
        # 기존 시트 삭제 (있다면)
        if site_code in workbook.sheetnames:
            del workbook[site_code]
        
        # 새 시트 생성
        new_ws = workbook.create_sheet(title=site_code)
        
        # 헤더 복사
        for col_idx, header in enumerate(headers, 1):
            new_ws.cell(row=1, column=col_idx, value=header)
            new_ws.cell(row=1, column=col_idx).fill = green_fill
        
        # 해당 사이트 데이터 필터링 및 복사
        filtered_df = df[df.iloc[:, 1].str.contains(site_filter, na=False)]
        
        for row_idx, (_, row) in enumerate(filtered_df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                new_ws.cell(row=row_idx, column=col_idx, value=value)
        
        # A열 순번 재설정
        for row_idx in range(2, new_ws.max_row + 1):
            new_ws.cell(row=row_idx, column=1, value=row_idx - 1)
        
        # 열 너비 복사
        for col_idx in range(1, len(headers) + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            new_ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width
    
    # 정렬 및 서식 적용
    center_alignment = Alignment(horizontal='center')
    right_alignment = Alignment(horizontal='right')
    
    # A, B열 가운데 정렬
    for col in ['A', 'B']:
        for cell in ws[col]:
            cell.alignment = center_alignment
    
    # D, E, G열 오른쪽 정렬
    for col in ['D', 'E', 'G']:
        for cell in ws[col]:
            cell.alignment = right_alignment
    
    # 첫 번째 행 가운데 정렬
    for cell in ws[1]:
        cell.alignment = center_alignment
    
    # 파일 저장
    output_path = file_path.replace('.xlsx', '_매크로_완료.xlsx')
    workbook.save(output_path)
    print(f"처리된 파일이 저장되었습니다: {output_path}")
    
    return output_path

def apply_conditional_formatting(file_path):
    """
    조건부 서식을 적용하는 함수 (색상, 굵기 등)
    """
    workbook = openpyxl.load_workbook(file_path)
    ws = workbook.active
    
    # 빨간색 폰트, 굵게
    red_font = Font(color="FF0000", bold=True)
    # 파란색 배경
    blue_fill = PatternFill(start_color="CCE8FF", end_color="CCE8FF", fill_type="solid")
    
    # 조건에 따른 서식 적용
    for row in range(2, ws.max_row + 1):
        # V열이 0인 경우 빨간색 굵게
        if ws[f'V{row}'].value == 0:
            ws[f'V{row}'].font = red_font
        
        # F열 정규식 패턴 매칭된 경우 파란색 배경
        f_val = str(ws[f'F{row}'].value)
        regex_exact = re.compile(r'^\d+개?$', re.IGNORECASE)
        regex_repeat = re.compile(r'(\d+개\s*){2,}', re.IGNORECASE)
        
        if regex_exact.match(f_val) or regex_repeat.search(f_val):
            ws[f'F{row}'].fill = blue_fill
        
        # 카카오 + 제주인 경우 F열 파란색 배경, J열 빨간색 굵게
        if ("카카오" in str(ws[f'B{row}'].value) and 
            "제주" in str(ws[f'J{row}'].value)):
            ws[f'F{row}'].fill = blue_fill
            ws[f'J{row}'].font = red_font
        
        # L열이 "착불"인 경우 빨간색 굵게
        if str(ws[f'L{row}'].value).strip() == "착불":
            ws[f'L{row}'].font = red_font
    
    workbook.save(file_path)
    return file_path

# 사용 예시
if __name__ == "__main__":
    # 파일 경로를 지정하세요
    excel_file_path = "./files/xlsx/test-[기본양식]-ERP용.xlsx"
    
    # 매크로 실행
    processed_file = other_site_macro_1_to_14(excel_file_path)
    
    # 조건부 서식 적용
    apply_conditional_formatting(processed_file)
    
    print("모든 처리가 완료되었습니다!")
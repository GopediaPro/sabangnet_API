import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from utils.excels.excel_handler import ExcelHandler


class ZigzagMacro:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.ex = ExcelHandler.from_file(file_path)
        self.ws = self.ex.ws
        self.wb = self.ex.wb
        self.last_row = self.ex.last_row
        self.last_col = self.ws.max_column
        self.headers = []
        self.df = None
        self.light_blue_fill = PatternFill(
            start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        self.green_fill = PatternFill(
            start_color="006100", end_color="006100", fill_type="solid")
        self.font = Font(name='맑은 고딕', size=9)
        self.center_alignment = Alignment(horizontal='center')
        self.right_alignment = Alignment(horizontal='right')

    def step_1_to_9(self) -> str:
        """
        1~9단계 자동화 실행
        """
        print("1~9단계 자동화 시작...")
        self._step_1()
        self._step_2()
        self._step_3()
        self._step_4()

        print("9단계: 모든 시트 서식 적용 시작...")
        for ws in self.wb.worksheets:
            if ws.max_row <= 1:
                continue
            self._step_9(ws)

        output_path = self.ex.save_file(self.file_path)
        print(f"✓ 지그재그 자동화 완료! 최종 파일: {output_path}")
        return output_path

    def _step_1(self):
        """
        [1단계] 데이터를 DataFrame으로 변환
        """
        # 데이터를 DataFrame으로 변환
        self.df = self.ex.to_dataframe(self.ws)
        self.headers = list(self.df.columns)

        # C열(인덱스 2), B열(인덱스 1) 순서로 정렬
        if len(self.df.columns) > 2:
            self.df = self.ex.sort_dataframe_by_c_b(
                self.df, c_col='수취인명', b_col='사이트')

        print("1단계: C열, B열 순서 정렬 완료")

    def _step_2(self):
        """
        [2단계] 워크시트에 정렬된 데이터 덮어쓰기
        """
        for row_idx, row_data in enumerate(self.df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                self.ws.cell(row=row_idx, column=col_idx,
                             value=value if value or value == 0 else "")
        print("2단계: 워크시트에 정렬된 데이터 덮어쓰기 완료")
    
    def _step_3(self):
        """
        [3단계] 시트 분리 준비
        """
        self.ws_map = self.ex.create_split_sheets(
            self.headers, ["OK", "IY"])

        self.ex.set_header_style(
            self.ws_map["OK"], self.headers, self.green_fill, self.font, self.center_alignment)
        self.ex.set_header_style(
            self.ws_map["IY"], self.headers, self.green_fill, self.font, self.center_alignment)

        print("3단계: 시트 분리 준비 완료")
    
    def _step_4(self):
        """
        [4단계] 시트 분리
        """
        site_mapping = {
            "OK": ["오케이마트"],
            "IY": ["아이예스"],
        }

        self.ex.split_sheets_by_site(self.df, self.ws_map, site_mapping)
        print("4단계: 시트 분리 완료")

    def _step_5(self):
        """
        [5단계] VLOOKUP 수식 입력 (V열)
        """
        last_row = self.ws.max_row
        for row in range(2, last_row + 1):
            self.ws[f'V{row}'].value = f'=VLOOKUP(M{row},Sheet1!$A:$B,2,0)'
        print("5단계: VLOOKUP 수식 입력 완료")

    def _step_6(self, ws):
        """
        [6단계] D열 수식 입력이나 복사
        """
        d2_formula = ws['D2'].value
        self.ex.autofill_d_column(
            start_row=2, end_row=ws.max_row, formula=d2_formula)
        print("6단계: D열 수식 입력 및 복사 완료")

    def _step_7(self,ws):
        """
        [6단계] F열 " 1개" 삭제 & M열 텍스트 열 생성자 변환
        """
        for row in range(2, ws.max_row + 1):
            m_value = ws[f'M{row}'].value
            f_value = ws[f'F{row}'].value
            if f_value is not None:
                ws[f'F{row}'].value = self.ex.clean_model_name(f_value)
            if m_value is not None:
                ws[f'M{row}'].value = str(m_value).strip()

        print("7단계: F열 모르겠는 셀 색칠음영 (하늘색) 완료")

    def _step_8(self):
        """
        [8단계] F열 모르겠는 셀 색칠음영 (하늘색)
        """
        last_row = self.ws.max_row
        self.ex.highlight_column(
            col='F', light_color=self.light_blue_fill, start_row=2, last_row=last_row)
        print("8단계: F열 모르겠는 셀 색칠음영 (하늘색) 완료")

    def _step_9(self,ws):
        """
        [9단계] A열 수식 값 변환
        """
        self.ex.set_row_number(ws)

        # A, B, D, E, G열 정렬
        self.ex.set_column_alignment(ws)
        
        # 테두리 제거 & 격자 제거
        self.ex.clear_borders(ws)

        # 색칠음영 제거
        self.ex.clear_fills_from_second_row()

        # 기본 폰트 적용
        self.ex.set_basic_format(ws=ws, header_rgb="008000")
        
        print(f"9단계: {ws.title} 시트에 서식 적용 완료")
        
        
        
        
        

import openpyxl
from utils.excels.excel_handler import ExcelHandler
from utils.excels.excel_column_handler import ExcelColumnHandler
from utils.macros.ERP.utils import average_duplicate_order_address_amounts
from utils.logs.sabangnet_logger import get_logger

logger = get_logger(__name__)

class ERPAliMacroV2:
    def __init__(self, file_path, is_star: bool = False):
        self.ex = ExcelHandler.from_file(file_path)
        self.file_path = file_path
        self.is_star = is_star
        self.ws = self.ex.ws
        self.wb = self.ex.wb

    def ali_erp_macro_run(self):
        """
        개선된 프로세스 순서:
        1. 시트 생성
        2. 데이터 처리 (컬럼 처리, VLOOKUP 딕셔너리 생성 등)
        3. 스타배송 모드에서 평균 금액 적용
        4. 시트별로 데이터 분리
        5. 시트별 디자인 적용
        """
        logger.info("=== 알리 ERP 자동화 V2 시작 ===")
        
        # 1단계: 시트 설정 및 생성
        sheets_name = ["OK", "IY"]
        site_to_sheet = {
            "오케이마트": "OK",
            "아이예스": "IY",
        }
        
        # 필요한 시트들이 없으면 생성
        self._ensure_sheets_exist(sheets_name)
        logger.info("✓ 시트 생성 완료")

        # 2단계: 데이터 처리
        logger.info("데이터 처리 시작...")
        col_h = ExcelColumnHandler()
        
        # 기본 데이터 처리
        for row in range(2, self.ws.max_row + 1):
            self._z_to_f_column(row)
            self._f_column_process(self.ws[f'F{row}'])
            self._i_to_h_column(self.ws[f'I{row}'], self.ws[f'H{row}'])
        
        # VLOOKUP 딕셔너리 생성
        vlookup_dict = self.ex.create_vlookup_dict(self.wb)
        logger.info("✓ VLOOKUP 딕셔너리 생성 완료")
        
        # D, U, V 컬럼 처리 (기본 데이터 처리 후)
        for row in range(2, self.ws.max_row + 1):
            col_h.d_column(
                self.ws[f'D{row}'], self.ws[f'U{row}'], self.ws[f'V{row}'])
        logger.info("✓ 기본 데이터 처리 완료")

        # 3단계: 스타배송 모드에서 평균 금액 적용
        if self.is_star:
            logger.info("스타배송 모드: 평균 금액 적용 중...")
            average_duplicate_order_address_amounts(self.ws)
            logger.info("✓ 평균 금액 적용 완료")

        # 4단계: 시트별로 데이터 분리
        logger.info("시트별 데이터 분리 시작...")
        sort_columns = [2, 3, 5]  # 정렬 기준
        headers, data = self.ex.preprocess_and_update_ws(self.ws, sort_columns)
        
        self.ex.split_and_write_ws_by_site(
            wb=self.wb,
            headers=headers,
            data=data,
            sheets_name=sheets_name,
            site_to_sheet=site_to_sheet,
            site_col_idx=2,
        )
        logger.info("✓ 시트별 데이터 분리 완료")

        # 5단계: 시트별 디자인 적용
        logger.info("시트별 서식, 디자인 적용 시작...")
        for ws in self.wb.worksheets:
            if ws.title == "Sheet":  # 기본 시트는 건너뛰기
                continue
                
            self.ex.set_header_style(ws)
            if ws.max_row <= 1:
                continue
                
            for row in range(2, ws.max_row + 1):
                if ws.title != "자동화":
                    col_h.a_value_column(ws[f"A{row}"])
                else:
                    col_h.a_formula_column(ws[f"A{row}"])
                col_h.convert_int_column(ws[f"P{row}"])
                col_h.convert_int_column(ws[f"Q{row}"])
                # VLOOKUP 적용
                self._vlookup_column(
                    ws[f"F{row}"], ws[f"S{row}"], vlookup_dict)
            
            logger.info(f"✓ [{ws.title}] 서식 및 디자인 적용 완료")

        # 최종 파일 저장
        output_path = self.ex.save_file(self.file_path)
        logger.info(f"✓ 알리 ERP 자동화 V2 완료! 최종 파일: {output_path}")
        return output_path

    def _ensure_sheets_exist(self, sheets_name):
        """
        필요한 시트들이 존재하는지 확인하고 없으면 생성
        """
        existing_sheets = [ws.title for ws in self.wb.worksheets]
        
        for sheet_name in sheets_name:
            if sheet_name not in existing_sheets:
                self.wb.create_sheet(title=sheet_name)
                logger.info(f"  - {sheet_name} 시트 생성됨")

    def _z_to_f_column(self, row):
        """
        Z열 -> F열 복사
        """
        self.ws[f'F{row}'].value = self.ws[f'Z{row}'].value

    def _f_column_process(self, cell):
        """
        F열 처리
        """
        if cell.value:
            txt = str(cell.value).strip()
            if txt.endswith(" * 1"):
                cell.value = txt[:-4]
            elif " * " in txt:
                parts = txt.split(" * ")
                if len(parts) >= 2:
                    suffix = parts[-1].strip()
                    if suffix.isdigit() and suffix != "1":
                        base_text = " * ".join(parts[:-1])
                        cell.value = f"{base_text} {suffix}개"

    def _i_to_h_column(self, i_cell, h_cell):
        """
        I열 -> H열 복사
        """
        if i_cell.value:
            ph_num = str(i_cell.value).replace("-", "").strip()
            if ph_num.isdigit():
                if len(ph_num) == 11:
                    formatted = self.ex.format_phone_number(ph_num)
                elif len(ph_num) in [9, 10]:
                    val = "010" + ph_num[-8:]
                    formatted = self.ex.format_phone_number(val)
                else:
                    formatted = i_cell.value
                i_cell.value = formatted
            h_cell.value = i_cell.value

    def _vlookup_column(self, key_cell, value_cell, vlookup_dict):
        """
        S열 VLOOKUP 및 값 붙여넣기 및 #N/A → "S"
        """
        # S열에 VLOOKUP 수식 입력 (임시로 "S" 값 설정)
        if vlookup_dict.get(str(key_cell.value)):
            value_cell.value = vlookup_dict.get(str(key_cell.value))
        else:
            value_cell.value = "S"

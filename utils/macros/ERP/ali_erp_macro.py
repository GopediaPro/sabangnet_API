import openpyxl
from utils.excels.excel_handler import ExcelHandler
from utils.excels.excel_column_handler import ExcelColumnHandler


class ERPAliMacro:
    def __init__(self, file_path):
        self.ex = ExcelHandler.from_file(file_path)
        self.file_path = file_path
        self.ws = self.ex.ws
        self.wb = self.ex.wb

    def ali_erp_macro_run(self):
        col_h = ExcelColumnHandler()

        for row in range(2, self.ws.max_row + 1):
            self._z_to_f_column(row)
            self._f_column_process(self.ws[f'F{row}'])
            self._i_to_h_column(self.ws[f'I{row}'], self.ws[f'H{row}'])
            col_h.d_column(
                # =U2+V2
                self.ws[f'D{row}'], self.ws[f'U{row}'], self.ws[f'V{row}'])

        sheets_name = ["OK", "IY"]
        site_to_sheet = {
            "오케이마트": "OK",
            "아이예스": "IY",
        }

        # 정렬 기준: 2번째 컬럼(B) → 3번째 컬럼(C) 순으로 정렬
        sort_columns = [2, 3, 5]
        print("시트별 정렬, 시트 분리 시작...")
        headers, data = self.ex.preprocess_and_update_ws(self.ws, sort_columns)
        self.ex.split_and_write_ws_by_site(
            wb=self.wb,
            headers=headers,
            data=data,
            sheets_name=sheets_name,
            site_to_sheet=site_to_sheet,
            site_col_idx=2,
        )
        print("시트별 정렬, 시트 분리 완료")

        print("시트별 서식, 디자인 적용 시작...")
        for ws in self.wb.worksheets:
            self.ex.set_header_style(ws)
            if ws.max_row <= 1:
                continue
            for row in range(2, ws.max_row + 1):
                if ws.title != "자동화":
                    col_h.a_value_column(ws[f"A{row}"])
                else:
                    col_h.a_formula_column(ws[f"A{row}"])
                self._jeju_address_column(ws, row, ws[f"J{row}"])
                col_h.convert_int_column(ws[f"P{row}"])
                col_h.convert_int_column(ws[f"Q{row}"])
            print(f"[{ws.title}] 서식 및 디자인 적용 완료")

        output_path = self.ex.save_file(self.file_path)
        print(f"✓ 알리 ERP 자동화 완료! 최종 파일: {output_path}")
        return output_path

    def _z_to_f_column(self, row):
        """
        Z열 -> F열 복사
        """
        self.ws[f'F{row}'].value = self.ws[f'Z{row}'].value

    def _f_column_process(self, cell):
        """
        F열 처리
        """
        if cell.value:
            txt = str(cell.value).strip()
            if txt.endswith(" * 1"):
                cell.value = txt[:-4]
            elif " * " in txt:
                parts = txt.split(" * ")
                if len(parts) >= 2:
                    suffix = parts[-1].strip()
                    if suffix.isdigit() and suffix != "1":
                        base_text = " * ".join(parts[:-1])
                        cell.value = f"{base_text} {suffix}개"

    def _i_to_h_column(self, i_cell, h_cell):
        """
        I열 -> H열 복사
        """
        if i_cell.value:
            ph_num = str(i_cell.value).replace("-", "").strip()
            if ph_num.isdigit():
                if len(ph_num) == 11:
                    formatted = self.ex.format_phone_number(ph_num)
                elif len(ph_num) in [9, 10]:
                    val = "010" + ph_num[-8:]
                    formatted = self.ex.format_phone_number(val)
                else:
                    formatted = i_cell.value
                i_cell.value = formatted
            h_cell.value = i_cell.value

    def _jeju_address_column(self, ws, row, cell):
        """
        제주 주소 포맷
        args:
            cell: 대상 셀 (J)
        """
        if cell.value and "제주" in str(cell.value):
            self.ex.process_jeju_address(
                ws, row, f_col='F', j_col='J')

    def _vlookup_column(self):
        """
        S열 VLOOKUP 및 값 붙여넣기 및 #N/A → "S"
        """
        # S열에 VLOOKUP 수식 입력 (임시로 "S" 값 설정)F
        for row in range(2, self.last_row + 1):
            # 실제 VLOOKUP 계산은 복잡하므로 기본값 "S" 설정
            self.ws[f'S{row}'].value = "S"
            self.ws[f'S{row}'].number_format = 'General'


def process_vlookup_simulation(file_path, lookup_sheet="sheet1", lookup_range="A:B"):
    """
    VLOOKUP 시뮬레이션 함수 (실제 조회 테이블이 있는 경우 사용)

    Args:
        file_path (str): Excel 파일 경로
        lookup_sheet (str): 조회할 시트명
        lookup_range (str): 조회 범위
    """
    try:
        workbook = openpyxl.load_workbook(file_path)

        # 조회 테이블이 있는지 확인
        if lookup_sheet in workbook.sheetnames:
            lookup_ws = workbook[lookup_sheet]

            # 실제 VLOOKUP 로직 구현
            for ws_name in workbook.sheetnames:
                ws = workbook[ws_name]
                if ws.max_row > 1:
                    for row in range(2, ws.max_row + 1):
                        f_value = ws[f'F{row}'].value
                        if f_value:
                            # 조회 로직 (간단한 예시)
                            lookup_result = "S"  # 기본값

                            # 실제 조회 테이블에서 값 찾기
                            for lookup_row in range(1, lookup_ws.max_row + 1):
                                lookup_key = lookup_ws[f'A{lookup_row}'].value
                                if lookup_key and str(lookup_key).strip() == str(f_value).strip():
                                    lookup_result = lookup_ws[f'B{lookup_row}'].value or "S"
                                    break

                            ws[f'S{row}'].value = lookup_result

            workbook.save(file_path)
            print("VLOOKUP 시뮬레이션 완료")
        else:
            print(f"조회 시트 '{lookup_sheet}'를 찾을 수 없습니다.")

    except Exception as e:
        print(f"VLOOKUP 처리 중 오류: {e}")

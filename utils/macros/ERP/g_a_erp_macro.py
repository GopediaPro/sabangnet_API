from openpyxl.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.logs.sabangnet_logger import get_logger
from utils.excels.excel_handler import ExcelHandler
from utils.excels.excel_column_handler import ExcelColumnHandler
from utils.macros.ERP.utils import average_duplicate_cart_address_amounts, average_duplicate_cart_address_amounts_df
import pandas as pd

logger = get_logger(__name__)

class ERPGmaAucMacro:
    def __init__(self, file_path: str, is_star: bool = False):
        self.ex: ExcelHandler = ExcelHandler.from_file(file_path)
        self.file_path = file_path
        self.is_star = is_star
        self.ws: Worksheet = self.ex.ws
        self.wb: Workbook = self.ex.wb
        self.basket_dict = {}
        self.headers: list[str] = None

    def gauc_erp_macro_run(self):

        # 1. 기존 시트 너비 가져오기
        sheet_widths: dict = self.ex.get_all_sheet_column_widths()

        df: pd.DataFrame = pd.read_excel(self.file_path)

        # 2. 기본 값처리
        self.run_df_gauc_erp_macro(df)
        logger.info("시트별 정렬, 기본 값 처리 완료")

        # 3. 스타배송 모드에서 장바구니번호-수취인주소 중복 금액 평균 처리
        if self.is_star:
            df = average_duplicate_cart_address_amounts_df(df)
            logger.info("장바구니번호-수취인주소 중복 금액 평균 처리 완료")

        # 4. 시트별 데이터 분리
        df_ok_cl_bb, df_iy = self._split_sheet_df(df)
        logger.info("시트별 데이터 분리 완료")

        # 4. 데이터 저장
        output_path = self.file_path.replace(".xlsx", "_매크로_완료.xlsx")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 데이터프레임들을 각 시트에 저장
            df.to_excel(writer, sheet_name="자동화", index=False)
            df_ok_cl_bb.to_excel(writer, sheet_name="OK,CL,BB", index=False)
            df_iy.to_excel(writer, sheet_name="IY", index=False)
            logger.info("시트별 데이터 적용 완료")

            logger.info("디자인 적용 시작...")
            wb = writer.book

            self._process_excel_style(wb, sheet_widths)

        logger.info(f"✓ G,옥 ERP 자동화 완료! 최종 파일: {output_path}")
        return output_path

    def run_df_gauc_erp_macro(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Dataframe으로 db데이터도 실행가능하게 분리
        """
        # 시트정렬
        df.sort_values(
            by=['사이트', '수취인명', '주문번호'],
            ascending=[True, True, False],
            inplace=True
        )

        # 중복 장바구니 처리
        self._process_dupl_basket(df)

        # A열 : 순번 수식으로 설정
        df['순번'] = "=ROW()-1"

        # D열 : 금액 계산
        df['금액'] = df['정산예정금액'] + df['서비스이용료'] + df['배송비']

        # F열 : 제품명에서 ' 1개' 제거
        df['제품명'] = df['제품명'].astype(str).str.replace(' 1개', '', regex=False)

        # L열 : 선/착불 처리
        df['선/착불'] = df['선/착불'].where(df['선/착불'] != "신용", "")

        return df

    def _process_dupl_basket(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        장바구니 중복값 배송비 제거
        """
        # 특정 키값 생성
        df['site_basket'] = df['사이트'] + '_' + df['장바구니번호'].astype(str)

        # 배송비 유효성 체크
        valid_delivery = (df['배송비'] != 0) & (
            df['배송비'].notna()) & (df['배송비'] != "")

        # 그룹별 첫 번째와 마지막 행 인덱스 찾기
        grouped = df.groupby('site_basket')
        first_idx = df[valid_delivery].groupby('site_basket').head(1).index
        last_idx = grouped.tail(1).index

        # 배송비 매핑 딕셔너리 생성
        delivery_mapping = df.loc[first_idx].set_index('site_basket')[
            '배송비'].to_dict()

        # 모든 배송비를 0으로 초기화
        df['배송비'] = 0

        # 마지막 행에만 배송비 적용
        mask = df.index.isin(last_idx)
        df.loc[mask, '배송비'] = df.loc[mask, 'site_basket'].map(
            delivery_mapping).fillna(0)

        df.drop('site_basket', axis=1, inplace=True)
        return df

    def _split_sheet_df(self, df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
        """
        시트 조건에 맞게 데이터 분리
        """
        # 사이트명 추출
        df['site_extracted'] = df['사이트'].str.extract(r'\[(.*?)\]')[0]

        # 조건별 마스크 생성
        ok_cl_bb_mask = df['site_extracted'].str.contains(
            '오케이마트|클로버프|베이지베이글', na=False)
        iy_mask = df['site_extracted'].str.contains('아이예스', na=False)

        # 데이터 분리
        df_ok_cl_bb = df[ok_cl_bb_mask].copy()
        df_iy = df[iy_mask].copy()

        # 임시 컬럼 제거
        for temp_df in [df, df_ok_cl_bb, df_iy]:
            if 'site_extracted' in temp_df.columns:
                temp_df.drop('site_extracted', axis=1, inplace=True)

        return df_ok_cl_bb, df_iy

    def _process_excel_style(self, wb, sheet_widths: dict):
        """
        엑셀 파일에 잘못된 값 표시와 디자인 적용 부분
        """
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
        col_h = ExcelColumnHandler()
        for ws in wb.worksheets:
            self.ex.set_header_style(ws)
            # 너비 적용
            for col in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = sheet_widths.get(
                    col_letter)
            for row in range(2, ws.max_row + 1):
                col_h.e_column(ws[f"E{row}"])
                col_h.f_column(ws[f"F{row}"])
                col_h.l_column(ws[f"L{row}"])
                col_h.convert_int_column(ws[f"P{row}"])
                col_h.convert_int_column(ws[f"R{row}"])
                col_h.convert_int_column(ws[f"S{row}"])
                col_h.convert_int_column(ws[f"V{row}"])
                # 각셀 포트 9로 적용
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(size=9)

            logger.info(f"[{ws.title}] 서식 및 디자인 적용 완료")

    def _add_basket_dict(self, basket_cell: Cell, shipping_cell: Cell, b_cell: Cell):
        """
        장바구니 중복 값 추가 (B셀과 Q셀 조합 기준)
        args:
            basket_cell: 장바구니 번호 셀 (Q)
            shipping_cell: 배송비 셀 (V)
            b_cell: B셀
        """
        basket_no = str(basket_cell.value).strip() if basket_cell.value else ""
        b_value = str(b_cell.value).strip() if b_cell.value else ""
        
        # B셀과 Q셀의 조합으로 키 생성
        combined_key = f"{b_value}_{basket_no}"

        if not basket_no or not b_value:
            return

        if combined_key not in self.basket_dict:
            if shipping_cell.value != 0 and shipping_cell.value != "":
                self.basket_dict[combined_key] = shipping_cell.value
                shipping_cell.value = 0
        else:
            shipping_cell.value = 0

    def _shipping_costs_column(self, basket_cell: Cell, shipping_cell: Cell, b_cell: Cell):
        """
        정렬된 데이터에서 B셀과 Q셀 조합 중복 값 중 첫 번째 값에 배송비 적용
        args:
            basket_cell: 장바구니 번호 셀 (Q)
            shipping_cell: 배송비 셀 (V)
            b_cell: B셀
        """
        basket_no = str(basket_cell.value).strip() if basket_cell.value else ""
        b_value = str(b_cell.value).strip() if b_cell.value else ""
        
        # B셀과 Q셀의 조합으로 키 생성
        combined_key = f"{b_value}_{basket_no}"

        if combined_key in self.basket_dict:
            shipping_cell.value = self.basket_dict[combined_key]
            del self.basket_dict[combined_key]

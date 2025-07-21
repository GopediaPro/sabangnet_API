"""기타사이트 합포장 자동화 모듈"""

from __future__ import annotations
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Set

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from utils.excel_handler import ExcelHandler

import pandas as pd

# 설정 상수
MALL_NAME = "기타사이트"
RED_FONT = Font(color="FF0000", bold=True)
BLUE_FILL = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")

# 시트 분리 설정 
ACCOUNT_MAPPING = {
    # TODO: 클로버프 확인 필요
    "OK": ["오케이마트"],
    "CL": ["클로버프"],
    "BB": ["베이지베이글"],
    "IY": ["아이예스"]
}

# 필수 생성 시트 목록
REQUIRED_SHEETS = list(ACCOUNT_MAPPING.keys())

# 사이트 설정
class ETCSiteConfig:
    # 배송비 분할 대상
    DELIVERY_SPLIT_SITES: Set[str] = {
        "롯데온", "보리보리", "스마트스토어", "톡스토어"
    }
    
    # 배송비 무료 사이트
    FREE_DELIVERY_SITES: Set[str] = {"오늘의집"}
    
    # 주문번호 길이 제한
    ORDER_NUMBER_LENGTHS: Dict[str, int] = {
        "YES24": 11,
        "CJ온스타일": 26,
        "GSSHOP": 21,
        "스마트스토어": 16,
        "에이블리": 13,
        "올웨이즈": 14,
        "위메프": 13,
        "인터파크": 12,
        "쿠팡": 13,
        "티몬": 12,
        "하이마트": 12
    }
    
    # 숫자 변환 대상 사이트
    NUMERIC_SITES: Set[str] = {
        "에이블리", "오늘의집", "쿠팡", "텐바이텐", "NS홈쇼핑", 
        "그립", "보리보리", "카카오선물하기", "톡스토어", "토스"
    }

class ETCOrderUtils:
    """주문번호 처리 유틸리티"""
    
    @staticmethod
    def clean_order_text(txt) -> str:
        """주문번호 문자열 정리 (안전한 타입 변환 포함)"""
        if not txt:
            return ""
        # 안전한 문자열 변환 (float, int, str 모두 처리)
        txt = str(txt).replace(" 1개", "").strip()
        txt = txt.replace("/", " + ")
        return txt

    @staticmethod
    def extract_bracket_text(text: str | None) -> str:
        """[계정명] 형식에서 계정명만 추출"""
        if not text:
            return ""
        match = re.search(r"\[(.*?)\]", str(text))
        return match.group(1) if match else ""

def process_order_numbers(ws: Worksheet) -> None:
    """
    🔄 ExcelHandler 후보
    사이트별 주문번호 처리
    """
    for row in range(2, ws.max_row + 1):
        site = str(ws[f"B{row}"].value or "")
        order_raw = str(ws[f"E{row}"].value or "")
        
        # 사이트별 주문번호 길이 제한 적용
        for site_name, length in ETCSiteConfig.ORDER_NUMBER_LENGTHS.items():
            if site_name in site:
                ws[f"E{row}"].value = order_raw[:length]
                break
                
        # 쿠팡 특수 처리
        if "쿠팡" in site and "/" in order_raw:
            slash_count = order_raw.count("/")
            pure_length = len(order_raw.replace("/", ""))
            each_len = pure_length // (slash_count + 1)
            ws[f"E{row}"].value = order_raw[:each_len]


def etc_site_merge_packaging(input_path: str) -> str:
    """기타사이트 주문 합포장 자동화 처리 (VBA 매크로 14단계 시트분리 포함)"""
    # Excel 파일 로드
    ex = ExcelHandler.from_file(input_path)
    ws = ex.ws
    
    # ========== VBA 매크로 단계: 원본 시트 자동화 처리 ==========
    # 원본 시트에 자동화 로직 적용 (행 분할 포함)
    splitter = ETCSheetManager(ws, ACCOUNT_MAPPING)
    splitter.apply_automation_logic(ws)
    
    # ========== VBA 매크로 14단계: 시트분리 (OK, CL, BB, IY) ==========
    # 자동화 처리가 완료된 원본 시트에서 시트분리 수행
    rows_by_sheet = splitter.get_rows_by_sheet()
    
    # 모든 필수 시트 생성 (데이터 유무와 무관하게 OK, CL, BB, IY 시트 생성)
    for sheet_name in REQUIRED_SHEETS:
        splitter.copy_to_new_sheet_simple(
            ex.wb,
            sheet_name, 
            rows_by_sheet.get(sheet_name, [])
        )
    
    # 저장
    base_name = Path(input_path).stem  # 확장자 제거한 파일명
    output_path = ex.happojang_save_file(base_name=base_name)
    ex.wb.close()
    
    return output_path


class ETCDeliveryFeeHandler:
    """사이트별 배송비 처리"""
    
    def __init__(self, ws: Worksheet):
        self.ws = ws
        
    def process_delivery_fee(self) -> None:
        """
        🔄 ExcelHandler 후보
        사이트별 배송비 처리
        - 롯데온/보리보리 등: 주문 수로 나누기
        - 오늘의집: 무료배송
        - 토스: 3만원 이상 무료
        - 분할된 행에 대한 재처리 포함
        """
        for row in range(2, self.ws.max_row + 1):
            site = str(self.ws[f"B{row}"].value or "")
            order_text = str(self.ws[f"X{row}"].value or "")
            v_cell = self.ws[f"V{row}"]
            
            # V열 값이 "/" 구분자 형태인 경우 첫 번째 값 사용
            v_value = v_cell.value
            if v_value and "/" in str(v_value):
                v_val = float(str(v_value).split("/")[0].strip() or 0)
            else:
                v_val = float(v_value or 0) if v_value else 0
            
            # U열 값도 동일하게 처리
            u_value = self.ws[f"U{row}"].value
            if u_value and "/" in str(u_value):
                u_val = float(str(u_value).split("/")[0].strip() or 0)
            else:
                u_val = float(u_value or 0) if u_value else 0

            # 분할된 행에서 V열이 비어있는 경우 배송비 로직 재적용
            if v_value is None or v_value == "":
                
                # 배송비 분할 대상
                if any(s in site for s in ETCSiteConfig.DELIVERY_SPLIT_SITES) and "/" in order_text:
                    count = len(order_text.split("/"))
                    # 기본 배송비 3000원 적용 후 분할
                    if count > 0:
                        v_cell.value = round(3000 / count)
                        v_cell.font = RED_FONT

                # 무료배송
                elif any(s in site for s in ETCSiteConfig.FREE_DELIVERY_SITES):
                    v_cell.value = 0
                    v_cell.font = RED_FONT

                # 토스 (3만원 이상 무료)
                elif "토스" in site:
                    v_cell.value = 0 if u_val > 30000 else 3000
                    v_cell.font = RED_FONT
                    
            else:
                # V열에 값이 있는 경우 기존 로직 적용
                
                # 배송비 분할 대상
                if any(s in site for s in ETCSiteConfig.DELIVERY_SPLIT_SITES) and "/" in order_text:
                    count = len(order_text.split("/"))
                    if v_val > 3000 and count > 0:
                        v_cell.value = round(v_val / count)
                        v_cell.font = RED_FONT

                # 무료배송
                elif any(s in site for s in ETCSiteConfig.FREE_DELIVERY_SITES):
                    v_cell.value = 0
                    v_cell.font = RED_FONT

                # 토스 (3만원 이상 무료)
                elif "토스" in site:
                    v_cell.value = 0 if u_val > 30000 else 3000
                    v_cell.font = RED_FONT


class ETCSpecialCaseHandler:
    """특수 케이스 처리"""
    
    def __init__(self, ws: Worksheet):
        self.ws = ws
        
    def process_kakao_jeju(self) -> None:
        """카카오 + 제주도 주문 처리"""
        for row in range(2, self.ws.max_row + 1):
            site = str(self.ws[f"B{row}"].value or "")
            addr = str(self.ws[f"J{row}"].value or "")
            
            if "카카오" in site and "제주" in addr:
                # F열 안내문구 추가 및 배경색
                f_cell = self.ws[f"F{row}"]
                if "[3000원 연락해야함]" not in str(f_cell.value):
                    f_cell.value = f"{f_cell.value} [3000원 연락해야함]"
                f_cell.fill = BLUE_FILL
                
                # J열 빨간색 굵게
                self.ws[f"J{row}"].font = RED_FONT
                
    def process_l_column(self) -> None:
        """L열 신용/착불 처리"""
        for row in range(2, self.ws.max_row + 1):
            val = str(self.ws[f"L{row}"].value or "")
            if val == "신용":
                self.ws[f"L{row}"].value = ""
            elif val == "착불":
                self.ws[f"L{row}"].font = RED_FONT


class ETCSheetManager:
    """기타사이트 시트 분리 및 자동화 로직 적용 (VBA 매크로 14단계 구현)"""
    
    def __init__(self, ws: Worksheet, account_mapping: Dict[str, List[str]]):
        self.ws = ws
        self.account_mapping = account_mapping
        self.last_row = ws.max_row
        self.last_col = ws.max_column
        
        # 열 너비 저장 (VBA 매크로와 동일)
        self.col_widths = [
            ws.column_dimensions[get_column_letter(c)].width
            for c in range(1, self.last_col + 1)
        ]

    def get_rows_by_sheet(self) -> Dict[str, List[int]]:
        """
        시트별 행 번호 매핑 생성 (VBA ExtractBracketText 로직 구현)
        자동화 처리가 완료된 원본 시트에서 현재 상태를 기준으로 매핑
        """
        rows_by_sheet = defaultdict(list)
        
        # 현재 시트의 실제 최대 행 수를 사용 (분할된 행 포함)
        current_max_row = self.ws.max_row
        
        for r in range(2, current_max_row + 1):
            # B열에서 [계정명] 추출 (VBA ExtractBracketText와 동일)
            account = ETCOrderUtils.extract_bracket_text(self.ws[f"B{r}"].value)
            
            # 각 시트의 계정 목록과 정확히 비교
            for sheet_name, accounts in self.account_mapping.items():
                if account in accounts:
                    rows_by_sheet[sheet_name].append(r)
                    break
                    
        return rows_by_sheet

    def create_empty_sheet(self, wb, sheet_name: str) -> Worksheet:
        """빈 시트 생성 (헤더와 열 너비만 복사) - VBA 매크로와 동일"""
        # 기존 시트 삭제 (VBA: On Error Resume Next)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            
        new_ws = wb.create_sheet(sheet_name)
        
        # 제목행 복사 (VBA: sourceSheet.Rows(1).Copy destSheet.Rows(1))
        for c in range(1, self.last_col + 1):
            new_ws.cell(row=1, column=c, 
                       value=self.ws.cell(row=1, column=c).value)
            # 열너비 복사 (VBA: destSheet.Columns(c).ColumnWidth = colWidth(c))
            new_ws.column_dimensions[get_column_letter(c)].width = self.col_widths[c - 1]
            
        return new_ws

    def copy_sheet_data(self, ws: Worksheet, row_indices: List[int]) -> None:
        """시트에 데이터 행 복사 (VBA: sourceSheet.Rows(r).Copy Destination:=destSheet.Rows(targetRow))"""
        if not row_indices:
            return
            
        target_row = 2
        for r in row_indices:
            for c in range(1, self.last_col + 1):
                ws.cell(row=target_row, column=c, 
                       value=self.ws.cell(row=r, column=c).value)
            target_row += 1
        
        # A열 순번 재부여 (VBA: destSheet.Cells(r, "A").Value = r - 1)
        for r in range(2, target_row):
            ws[f"A{r}"].value = r - 1

    def copy_to_new_sheet_simple(self, 
                                wb, 
                                sheet_name: str, 
                                row_indices: List[int] = None) -> None:
        """
        VBA와 동일한 방식으로 새 시트 생성
        - 원본 시트의 처리된 데이터를 각 계정명별로 복사만 수행
        - 추가 자동화 로직 적용하지 않음 (VBA 매크로와 동일)
        """
        new_ws = self.create_empty_sheet(wb, sheet_name)
        if row_indices:
            self.copy_sheet_data_simple(new_ws, row_indices)
    
    def copy_sheet_data_simple(self, ws: Worksheet, row_indices: List[int]) -> None:
        """
        VBA와 동일한 방식으로 시트에 데이터 행 복사
        - 원본 시트의 현재 상태(분할된 행 포함)를 그대로 복사
        - A열 순번만 재부여
        """
        if not row_indices:
            return
            
        target_row = 2
        for r in row_indices:
            for c in range(1, self.last_col + 1):
                original_value = self.ws.cell(row=r, column=c).value
                ws.cell(row=target_row, column=c, value=original_value)
                
                # 원본 시트의 셀 서식도 복사
                original_cell = self.ws.cell(row=r, column=c)
                new_cell = ws.cell(row=target_row, column=c)
                
                # 폰트 복사
                if original_cell.font:
                    new_cell.font = Font(
                        color=original_cell.font.color,
                        bold=original_cell.font.bold,
                        name=original_cell.font.name,
                        size=original_cell.font.size
                    )
                
                # 배경색 복사
                if original_cell.fill:
                    new_cell.fill = PatternFill(
                        start_color=original_cell.fill.start_color,
                        end_color=original_cell.fill.end_color,
                        fill_type=original_cell.fill.fill_type
                    )
                
                # 정렬 복사
                if original_cell.alignment:
                    new_cell.alignment = Alignment(
                        horizontal=original_cell.alignment.horizontal,
                        vertical=original_cell.alignment.vertical
                    )
            
            target_row += 1
        
        # 데이터가 2행 이상 있는 경우 정렬 수행
        if target_row > 3:  # 헤더(1행) + 데이터(2행 이상)
            ex = ExcelHandler(ws)
            ex.sort_by_columns([2, 3])  # B열, C열 기준 정렬
        
        # A열 순번 재부여 "=ROW()-1"
        for r in range(2, target_row):
            ws[f"A{r}"].value = "=ROW()-1"

    def apply_automation_logic(self, ws: Worksheet) -> None:
        """자동화 로직 적용 (VBA 매크로 21단계까지의 모든 로직)"""
        # Excel 핸들러로 기본 처리 적용
        ex = ExcelHandler(ws)
        
        # 1. 기본 서식 적용
        ex.set_basic_format()
        
        # 2. C→B 정렬
        ex.sort_by_columns([2, 3])
        
        # 3. D열 수식 설정 (사용자 요구사항에 맞는 P, V 처리)
        self._calculate_d_column_custom(ws)
        
        # 4. 사이트별 배송비 처리
        ETCDeliveryFeeHandler(ws).process_delivery_fee()
        
        # 5. 주문번호 처리
        process_order_numbers(ws)
        
        # 6. 전화번호 처리
        for row in range(2, ws.max_row + 1):
            for col in ('H', 'I'):
                cell_value = ws[f'{col}{row}'].value
                ws[f'{col}{row}'].value = ex.format_phone_number(cell_value)
        
        # 7. 특수 케이스 처리
        special = ETCSpecialCaseHandler(ws)
        special.process_kakao_jeju()
        special.process_l_column()
        
        # 8. F열 텍스트 정리
        for row in range(2, ws.max_row + 1):
            ws[f"F{row}"].value = ETCOrderUtils.clean_order_text(ws[f"F{row}"].value)

        # 9. 문자열→숫자 변환 
        ex.convert_numeric_strings(cols=("F","M", "W", "AA"))

        # 10. 열 정렬
        ex.set_column_alignment()
        # F열 왼쪽정렬 
        for row in range(1, ws.max_row + 1):
            ws[f"F{row}"].alignment = Alignment(horizontal='left')

        # 11. 배경·테두리 제거, A열 순번 설정
        self.set_row_number(ws)  # 자체 정의한 메서드 사용
        ex.clear_fills_from_second_row()
        ex.clear_borders()

    def _calculate_d_column_custom(self, ws: Worksheet) -> None:
        """
        D열 계산: O + P(슬래시 합산) + V(슬래시 합산)
        - P열: "780/780" → 780 + 780 = 1560
        - V열: "3000/3000" → 3000 + 3000 = 6000
        """
        for row in range(2, ws.max_row + 1):
            o_val = ws[f'O{row}'].value or 0
            p_val = ws[f'P{row}'].value or 0
            v_val = ws[f'V{row}'].value or 0
            
            # O열 처리 (숫자로 변환)
            try:
                o_num = float(o_val) if o_val else 0
            except (ValueError, TypeError):
                o_num = 0
            
            # P열 처리: "/" 구분자가 있으면 모든 숫자를 합산
            p_num = 0
            if p_val and "/" in str(p_val):
                p_parts = str(p_val).split("/")
                for part in p_parts:
                    try:
                        p_num += float(part.strip())
                    except (ValueError, TypeError):
                        pass  # 숫자가 아닌 경우 무시
            else:
                try:
                    p_num = float(p_val) if p_val else 0
                except (ValueError, TypeError):
                    p_num = 0
            
            # V열 처리: "/" 구분자가 있으면 모든 숫자를 합산 (P열과 동일)
            v_num = 0
            if v_val and "/" in str(v_val):
                v_parts = str(v_val).split("/")
                for part in v_parts:
                    try:
                        v_num += float(part.strip())
                    except (ValueError, TypeError):
                        pass  # 숫자가 아닌 경우 무시
            else:
                try:
                    v_num = float(v_val) if v_val else 0
                except (ValueError, TypeError):
                    v_num = 0
            
            # D열에 계산 결과 설정
            calculated_d = o_num + p_num + v_num
            ws[f'D{row}'].value = calculated_d

    def _calculate_d_column(self, ws: Worksheet, row: int) -> None:
        """D열 계산: O + P + V (slash 처리 후 첫 값 사용)"""
        o_val = ws[f'O{row}'].value or 0
        p_val = ws[f'P{row}'].value or 0
        v_val = ws[f'V{row}'].value or 0
        
        # P열이 "/" 구분자 포함 시 첫 번째 값만 사용
        if p_val and "/" in str(p_val):
            p_val = float(str(p_val).split("/")[0].strip() or 0)
        else:
            p_val = float(p_val) if p_val else 0
        
        # V열이 "/" 구분자 포함 시 첫 번째 값만 사용
        if v_val and "/" in str(v_val):
            v_val = float(str(v_val).split("/")[0].strip() or 0)
        else:
            v_val = float(v_val) if v_val else 0
        
        try:
            calculated_d = float(o_val) + p_val + v_val
            ws[f'D{row}'].value = calculated_d
        except (ValueError, TypeError):
            ws[f'D{row}'].value = ""

    def _split_slash_values(self, value, expected_count: int) -> List[str]:
        """
        "/" 구분자로 나뉜 값들을 리스트로 분할
        expected_count만큼 값이 없으면 빈 문자열로 채움
        """
        if not value:
            return [""] * expected_count
        
        val_str = str(value).strip()
        if "/" in val_str:
            parts = val_str.split("/")
            # 파트가 expected_count보다 적으면 빈 문자열로 채움
            while len(parts) < expected_count:
                parts.append("")
            result = [part.strip() for part in parts[:expected_count]]
        else:
            # "/" 구분자가 없으면 첫 번째 값만 사용하고 나머지는 빈 문자열
            result = [val_str] + [""] * (expected_count - 1)
        
        return result

    def copy_to_new_sheet(self, 
                         wb, 
                         sheet_name: str, 
                         row_indices: List[int] = None) -> None:
        """지정된 행들로 새 시트 생성 (데이터가 없어도 빈 시트 생성)"""
        new_ws = self.create_empty_sheet(wb, sheet_name)
        if row_indices:
            self.copy_sheet_data(new_ws, row_indices)
            
        # 모든 시트에 자동화 로직 적용
        self.apply_automation_logic(new_ws)

    def set_row_number(self, ws: Worksheet, start_row: int = 2) -> None:
        """
        A열 순번 자동 생성 (=ROW()-1) - 현재 시트의 실제 행 수를 기준으로 처리
        분할된 행들을 포함한 모든 행에 순번을 정확히 설정
        """
        # 현재 시트의 실제 최대 행 수를 동적으로 확인
        end_row = ws.max_row
        
        for row in range(start_row, end_row + 1):
            ws[f'A{row}'].number_format = 'General'
            ws[f"A{row}"].value = "=ROW()-1"

if __name__ == "__main__":
    excel_file_path = "/Users/smith/Documents/github/OKMart/sabangnet_API/files/test-[기본양식]-합포장용.xlsx"
    processed_file = etc_site_merge_packaging(excel_file_path)
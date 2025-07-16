"""브랜디 합포장 자동화 모듈"""

from __future__ import annotations
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment

from utils.excel_handler import ExcelHandler


# 설정 상수
MALL_NAME = "브랜디"
RED_FONT = Font(color="FF0000", bold=True)
FONT_MALGUN = Font(name="맑은 고딕", size=9)
HDR_FILL = PatternFill(start_color="006100",
                       end_color="006100", fill_type="solid")
BLUE_FILL = PatternFill(start_color="CCE8FF", end_color="CCE8FF", fill_type="solid")
NO_BORDER = Border()
MULTI_SEP_RE = re.compile(r"[\/;]")

class BrandyProductProcessor:
    """브랜디 상품 정보 처리 유틸리티"""
    
    @staticmethod
    def clean_product_text(txt: str | None) -> str:
        """
        🔄 ExcelHandler 후보
        상품명 문자열 정리 (' 1개' 제거)
        """
        return str(txt or "").replace(" 1개", "").strip()


class BrandyPhoneFormatter:
    """브랜디 전화번호 처리 유틸리티"""
    
    @staticmethod
    def format_phone(val: str | None) -> str:
        """전화번호 포맷팅 (01012345678 → 010-1234-5678)"""
        if not val:
            return ""
        digits = re.sub(r"\D", "", str(val))
        if len(digits) == 11:
            return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
        return str(val)


class BrandyOrderMerger:
    """브랜디 주문 데이터 그룹핑 및 병합 처리"""
    
    def __init__(self, ws: Worksheet):
        self.ws = ws
        self.groups = defaultdict(list)
        
    def group_by_product_and_receiver(self) -> None:
        """C열(상품번호)와 J열(수령인) 기준으로 그룹핑"""
        for row in range(2, self.ws.max_row + 1):
            key = (
                f"{str(self.ws[f'C{row}'].value).strip()}"
                f"|{str(self.ws[f'J{row}'].value).strip()}"
            )
            self.groups[key].append(row)
            
    def merge_rows(self) -> List[int]:
        """그룹별 데이터 병합 처리"""
        rows_to_delete = []
        
        for rows in self.groups.values():
            if len(rows) == 1:  # 중복 없음
                continue
                
            base_row = rows[0]  # 첫 행 유지
            
            # D열 금액 합산
            total_d = 0.0
            for row in rows:
                cell_val = self.ws[f"D{row}"].value
                if isinstance(cell_val, str) and cell_val.startswith("="):
                    # 수식이 있는 경우 O+P+V 각각 계산
                    o_val = float(self.ws[f"O{row}"].value or 0)
                    p_val = float(self.ws[f"P{row}"].value or 0)
                    v_val = float(self.ws[f"V{row}"].value or 0)
                    total_d += (o_val + p_val + v_val)
                else:
                    total_d += float(cell_val or 0)
            self.ws[f"D{base_row}"].value = total_d
            
            # G열 수량 합산 2025-07-16 수량은 합산처리 대상 제외
            # total_g = 0
            # for row in rows:
            #     g_val = self.ws[f"G{row}"].value
            #     if g_val is not None:
            #         try:
            #             total_g += float(str(g_val).strip() or 0)
            #         except ValueError:
            #             pass  # 숫자로 변환할 수 없는 경우 무시
            # self.ws[f"G{base_row}"].value = total_g
            
            # F열 모델명 결합
            models = []
            for row in rows:
                model = self.ws[f"F{row}"].value
                if model:
                    clean_model = BrandyProductProcessor.clean_product_text(model)
                    if clean_model:
                        models.append(clean_model)
            self.ws[f"F{base_row}"].value = " + ".join(models)
            
            # 나머지 행은 삭제 대상으로 표시
            rows_to_delete.extend(rows[1:])
            
        return sorted(rows_to_delete, reverse=True)  # 역순 정렬(삭제용)

class BrandySheetProcessor:
    """브랜디 시트 분리 및 자동화 로직 적용"""
    
    def __init__(self, ws: Worksheet):
        self.ws = ws
        self.last_row = ws.max_row
        self.last_col = ws.max_column
        
        # 열 너비 저장
        self.col_widths = [
            ws.column_dimensions[get_column_letter(c)].width
            for c in range(1, self.last_col + 1)
        ]

    def create_empty_sheet(self, wb: Worksheet, sheet_name: str) -> Worksheet:
        """빈 시트 생성 (헤더와 열 너비만 복사)"""
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            
        new_ws = wb.create_sheet(sheet_name)
        
        # 헤더와 열 너비 복사
        for c in range(1, self.last_col + 1):
            new_ws.cell(row=1, column=c, 
                       value=self.ws.cell(row=1, column=c).value)
            new_ws.column_dimensions[get_column_letter(c)].width = self.col_widths[c - 1]
            
        return new_ws

    def copy_sheet_data(self, ws: Worksheet) -> None:
        """시트에 데이터 행 복사"""
        for r in range(2, self.last_row + 1):
            for c in range(1, self.last_col + 1):
                ws.cell(row=r, column=c, 
                       value=self.ws.cell(row=r, column=c).value)
            ws[f"A{r}"].value = "=ROW()-1"

    def apply_automation_logic(self, ws: Worksheet) -> None:
        """자동화 로직 적용"""
        # 1. 기본 서식 적용
        ex = ExcelHandler(ws)
        ex.set_basic_format()
        
        # 2. P, V열 "/" 구분자 합산 처리
        self.process_p_column_slash_values(ws)
        self.process_v_column_slash_values(ws)
        
        # 3. D열에 O+P+V 값 계산하여 입력
        self.calculate_d_column_values(ws)


        # 4. D열 기준 숫자 오름차순 정렬
        self.sort_by_d_column_numeric(ws)
        
        # 5. 그룹핑 및 병합
        merger = BrandyOrderMerger(ws)
        merger.group_by_product_and_receiver()
        rows_to_delete = merger.merge_rows()

        # 6. F열 모델명 정리 (모든 행에 대해 "1개" 제거)
        left_alignment = Alignment(horizontal='left')
        
        for row in range(2, ws.max_row + 1):
            model_value = ws[f'F{row}'].value
            if model_value:
                ws[f'F{row}'].value = BrandyProductProcessor.clean_product_text(model_value)
            # F열 왼쪽 정렬 적용
            ws[f'F{row}'].alignment = left_alignment
        
        # 중복 행 삭제 (역순으로)
        for row_idx in rows_to_delete:
            ws.delete_rows(row_idx)

        # 7. A열 순번 재설정
        ex.set_row_number(ws)
        
        # 8. 전화번호 처리 (H열, I열)
        for row in range(2, self.last_row + 1):
            for col in ('H', 'I'):
                cell_value = ws[f'{col}{row}'].value
                ws[f'{col}{row}'].value = ex.format_phone_number(cell_value)
        
        # 9. 제주도 주문 처리
        for row in range(2, self.last_row + 1):
            j_value = ws[f'J{row}'].value
            if j_value and "제주" in str(j_value):
                ex.process_jeju_address(row)

        # 10. 문자열→숫자 변환 2025-07-16 숫자처리 대상 조정
        ex.convert_numeric_strings(cols=("D", "O", "P", "U", "V"))
        # H열 왼쪽정렬 
        for row in range(1, ws.max_row + 1):
            ws[f"H{row}"].alignment = Alignment(horizontal='left')
        # I열 왼쪽정렬 
        for row in range(1, ws.max_row + 1):
            ws[f"I{row}"].alignment = Alignment(horizontal='left')
        # Q열 왼쪽정렬 
        for row in range(1, ws.max_row + 1):
            ws[f"Q{row}"].alignment = Alignment(horizontal='left')

        # 11. 열 정렬
        ex.set_column_alignment()
        
        # 12. 배경·테두리 제거
        ex.clear_fills_from_second_row()
        ex.clear_borders()

    def copy_to_new_sheet(self, 
                         wb: Worksheet, 
                         sheet_name: str,
                         ex: ExcelHandler) -> None:
        """새 시트 생성 및 자동화 로직 적용"""
        new_ws = self.create_empty_sheet(wb, sheet_name)
        self.copy_sheet_data(new_ws)
        self.apply_automation_logic(new_ws)  # ex 인자 제거

    def calculate_d_column_values(self, ws: Worksheet) -> None:
        """
        D열에 O+P+V 합계값 계산하여 직접 입력 (수식이 아닌 값)
        """
        for row in range(2, ws.max_row + 1):
            # O, P, V 열 값 읽기
            o_val = float(ws[f'O{row}'].value or 0)
            p_val = float(ws[f'P{row}'].value or 0)
            v_val = float(ws[f'V{row}'].value or 0)
            
            # D열에 합계 값 입력
            ws[f'D{row}'].value = o_val + p_val + v_val
            ws[f'D{row}'].number_format = 'General'

    def process_p_column_slash_values(self, ws: Worksheet) -> None:
        """
        P열의 "/" 구분자로 나뉜 숫자들의 합계 계산 
        예: "2600/308" → 2908
        """
        for row in range(2, ws.max_row + 1):
            cell_val = ws[f'P{row}'].value
            if cell_val and "/" in str(cell_val):
                parts = str(cell_val).split("/")
                total = 0.0
                for part in parts:
                    part = part.strip()
                    if part and part.replace('.', '').replace('-', '').isdigit():
                        try:
                            total += float(part)
                        except ValueError:
                            continue
                ws[f'P{row}'].value = total
        
    def process_v_column_slash_values(self, ws: Worksheet) -> None:
        """
        P열의 "/" 구분자로 나뉜 숫자들의 합계 계산 
        예: "2600/308" → 2908
        """
        for row in range(2, ws.max_row + 1):
            cell_val = ws[f'V{row}'].value
            if cell_val and "/" in str(cell_val):
                parts = str(cell_val).split("/")
                total = 0.0
                for part in parts:
                    part = part.strip()
                    if part and part.replace('.', '').replace('-', '').isdigit():
                        try:
                            total += float(part)
                        except ValueError:
                            continue
                ws[f'V{row}'].value = total

    def sort_by_d_column_numeric(self, ws: Worksheet) -> None:
        """
        D열을 숫자 기준으로 오름차순 정렬
        문자열 정렬이 아닌 실제 숫자값 기준으로 정렬
        """
        # 데이터를 읽어서 (행번호, 전체행데이터, D열숫자값) 튜플로 구성
        data_rows = []
        for row in range(2, ws.max_row + 1):
            row_data = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
            d_value = ws.cell(row=row, column=4).value  # D열 (4번째 열)
            
            # D열 값을 숫자로 변환 시도
            try:
                d_numeric = float(d_value) if d_value is not None else 0.0
            except (ValueError, TypeError):
                d_numeric = 0.0
                
            data_rows.append((row, row_data, d_numeric))
        
        # D열 숫자값 기준으로 오름차순 정렬
        data_rows.sort(key=lambda x: x[2])
        
        # 정렬된 데이터를 다시 시트에 쓰기
        for idx, (original_row, row_data, d_numeric) in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                # 2025-07-16 엑셀에서 None이 입력되면 이전 값으로 처리되어 빈값으로 대치
                if value is None:
                    value = ""
                ws.cell(row=idx, column=col_idx, value=value)

def brandy_merge_packaging(input_path: str) -> str:
    """브랜디 주문 합포장 자동화 처리"""
    # Excel 파일 로드
    ex = ExcelHandler.from_file(input_path)
    
    # 첫 번째 시트에 자동화 로직 적용
    source_ws = ex.ws
    splitter = BrandySheetProcessor(source_ws)
    splitter.apply_automation_logic(source_ws)
    print(f"◼︎ [{MALL_NAME}] 자동화 처리 완료")
    
    # 저장
    base_name = Path(input_path).stem  # 확장자 제거한 파일명
    output_path = ex.happojang_save_file(base_name=base_name)
    ex.wb.close()
    
    print(f"◼︎ [{MALL_NAME}] 합포장 자동화 완료!")
    return output_path


if __name__ == "__main__":
    test_path = "/Users/smith/Documents/github/OKMart/sabangnet_API/files/test-[기본양식]-합포장용.xlsx"
    brandy_merge_packaging(test_path)
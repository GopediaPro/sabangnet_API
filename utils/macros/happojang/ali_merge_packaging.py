"""알리익스프레스 합포장 자동화 모듈"""

from __future__ import annotations
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from utils.excel_handler import ExcelHandler


# 설정 상수
MALL_NAME = "알리익스프레스"
RED_FONT = Font(color="FF0000", bold=True)


class ProductUtils:
    # 정규식 패턴
    STAR_QTY_RE = re.compile(r"\* ?(\d+)")
    MULTI_SEP_RE = re.compile(r"[\/;]")
    DUP_QTY_RE = re.compile(r"\d+개")
    PHONE_RE = re.compile(r"\D")
    JEJU_RE = re.compile(r"제주|서귀포")
    
    @staticmethod
    def clean_product_text(txt: str | None) -> str:
        """
        🔄 ExcelHandler 후보
        상품명 문자열 정리
        - '/' ';' → ' + '
        - '*n' → 'n개' (n>1)
        - ' 1개' 제거
        """
        if not txt:
            return ""
            
        txt = ProductUtils.MULTI_SEP_RE.sub(" + ", str(txt))
        
        def qty_replace(m: re.Match) -> str:
            n = m.group(1).strip()
            return "" if n == "1" else f" {n}개"
            
        txt = ProductUtils.STAR_QTY_RE.sub(qty_replace, txt)
        return txt.replace(" 1개", "").strip()

    @staticmethod
    def format_phone(val: str | None) -> str:
        """전화번호 포맷 (01012345678 → 010-1234-5678)"""
        if not val:
            return ""
        digits = ProductUtils.PHONE_RE.sub("", str(val))
        if digits.startswith("10"):
            digits = "0" + digits
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}" if len(digits) == 11 else digits

    @staticmethod
    def check_multiple_quantities(txt: str) -> bool:
        """F열 다중수량 체크 ('개' 2회 이상 등장)"""
        parts = [p.strip() for p in str(txt or "").split("+")]
        return sum(1 for p in parts if ProductUtils.DUP_QTY_RE.search(p)) >= 2

    @staticmethod
    def is_jeju_address(addr: str) -> bool:
        """주소가 제주도인지 확인"""
        return bool(ProductUtils.JEJU_RE.search(str(addr or "")))

    @staticmethod
    def build_lookup_map(ws_lookup: Worksheet) -> Dict[str, str]:
        """Sheet1의 A:B를 딕셔너리로 변환"""
        return {
            str(r[0]): r[1]
            for r in ws_lookup.iter_rows(min_row=2, max_col=2, values_only=True)
            if r[0] is not None
        }


FONT_MALGUN = Font(name="맑은 고딕", size=9)
HDR_FILL = PatternFill(start_color="006100",
                       end_color="006100", fill_type="solid")
BLUE_FILL = PatternFill(start_color="CCE8FF",
                        end_color="CCE8FF", fill_type="solid")
JEJU_FILL = PatternFill(start_color="DDEBF7",
                        end_color="DDEBF7", fill_type="solid")
NO_BORDER = Border()


def to_num(val) -> float:
    """'12,345원' → 12345.0 (실패 시 0)."""
    try:
        return float(re.sub(r"[^\d.-]", "", str(val))) if str(val).strip() else 0.0
    except ValueError:
        return 0.0


def copy_product_info(ws: Worksheet) -> None:
    """Z열 상품정보를 F열로 복사하고 정리"""
    for row in range(2, ws.max_row + 1):
        z_val = ws[f"Z{row}"].value
        ws[f"F{row}"].value = ProductUtils.clean_product_text(z_val)
        if ProductUtils.check_multiple_quantities(ws[f"F{row}"].value):
            ws[f"F{row}"].fill = BLUE_FILL


def process_phones(ws: Worksheet) -> None:
    """전화번호 처리 (I열 포맷 + H열 복사)"""
    for row in range(2, ws.max_row + 1):
        phone = ProductUtils.format_phone(ws[f"I{row}"].value)
        ws[f"I{row}"].value = phone
        ws[f"H{row}"].value = phone
    ws.column_dimensions["H"].width = ws.column_dimensions["I"].width


def process_jeju_orders(ex: ExcelHandler) -> None:
    """제주도 주문 처리"""
    ws = ex.ws
    for row in range(2, ws.max_row + 1):
        if ProductUtils.is_jeju_address(ws[f"J{row}"].value):
            ws[f"J{row}"].font = RED_FONT
            if "[3000원 환불처리]" not in str(ws[f"F{row}"].value):
                ws[f"F{row}"].value = f"{ws[f'F{row}'].value} [3000원 환불처리]"
            ws[f"F{row}"].fill = BLUE_FILL


def ali_merge_packaging(input_path: str) -> str:
    """알리익스프레스 주문 합포장 자동화 처리"""
    # Excel 파일 로드
    ex = ExcelHandler.from_file(input_path)
    ws = ex.ws

    # 1. 기본 서식 적용
    ex.set_basic_format()
    
    # 2. P열 슬래시(/) 금액 합산
    ex.sum_prow_with_slash()
    
    # 3. C→B 정렬
    ex.sort_by_columns([3, 2])  # C열=3, B열=2
    
    # 4. 상품정보 복사 및 정리 (Z → F)
    copy_product_info(ws)
    
    # 5. 배경 제거
    ex.clear_fills_from_second_row()
    
    # 7-8. 전화번호 처리
    process_phones(ws)
    
    # 9-10. E열 LEFT(16) 처리
    ws.insert_cols(6)
    ws["F1"].value = "TrimE"
    for r in range(2, ws.max_row + 1):
        ws[f"F{r}"].value = str(ws[f"E{r}"].value)[:16]
    for r in range(2, ws.max_row + 1):
        ws[f"E{r}"].value = ws[f"F{r}"].value
    ws.delete_cols(6)
    
    # 12. D열 수식 설정
    ex.autofill_d_column(formula="=U{row}+P{row}+V{row}")
    
    # 13. A열 순번 설정
    ex.set_row_number(ws)
    
    # 14-17. 열 정렬/서식
    ex.set_column_alignment()
    ex.clear_borders()
    ws.column_dimensions["E"].width = 20
    
    # 19. 제주도 주문 처리
    process_jeju_orders(ex)
    
    # E, M, P, Q, W 열 String숫자 to 숫자 변환
    ex.convert_numeric_strings(cols=("E", "M", "P", "Q", "W"))
    
    # 21. VLOOKUP 처리 (Sheet1이 있는 경우)
    if "Sheet1" in ex.wb.sheetnames:
        lookup_map = ProductUtils.build_lookup_map(ex.wb["Sheet1"])
        for row in range(2, ws.max_row + 1):
            m_val = str(ws[f"M{row}"].value)
            ws[f"V{row}"].value = lookup_map.get(m_val, "S")
    
    # 22. 시트 분리 (OK, IY)
    splitter = SheetSplitter(ws)
    rows_by_sheet = splitter.get_rows_by_sheet()
    
    for sheet_name, row_indices in rows_by_sheet.items():
        if row_indices:  # 해당 사이트의 데이터가 있는 경우만
            splitter.copy_to_new_sheet(ex.wb, sheet_name, row_indices)

    # 24. 시트 순서 정리
    desired = ["알리합포장", "OK", "IY", "Sheet1"]
    for name in reversed(desired):
        if name in ex.wb.sheetnames:
            ex.wb._sheets.insert(0, ex.wb._sheets.pop(ex.wb.sheetnames.index(name)))
            
    # 저장
    base_name = Path(input_path).stem  # 확장자 제거한 파일명
    output_path = ex.happojang_save_file(base_name=base_name)
    
    print(f"◼︎ [{MALL_NAME}] 합포장 자동화 완료!")
    
    return output_path


class SheetSplitter:
    """시트 분리 처리 클래스"""
    
    def __init__(self, ws: Worksheet):
        self.ws = ws
        self.last_row = ws.max_row
        self.last_col = ws.max_column
        
        # 열 너비 저장
        self.col_widths = [
            ws.column_dimensions[get_column_letter(c)].width
            for c in range(1, self.last_col + 1)
        ]

    def get_rows_by_sheet(self) -> Dict[str, List[int]]:
        """사이트별 행 번호 매핑 생성"""
        site_rows = defaultdict(list)
        for r in range(2, self.last_row + 1):
            text = str(self.ws[f"B{r}"].value or "")
            if "오케이마트" in text:
                site_rows["OK"].append(r)
            elif "아이예스" in text:
                site_rows["IY"].append(r)
        return site_rows

    def copy_to_new_sheet(self, 
                         wb: Workbook, 
                         sheet_name: str, 
                         row_indices: List[int]) -> None:
        """지정된 행들을 새 시트로 복사"""
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            
        new_ws = wb.create_sheet(sheet_name)
        
        # 헤더와 열 너비 복사
        for c in range(1, self.last_col + 1):
            new_ws.cell(row=1, column=c, 
                       value=self.ws.cell(row=1, column=c).value)
            new_ws.column_dimensions[get_column_letter(c)].width = self.col_widths[c - 1]
        
        # 데이터 복사
        for idx, r in enumerate(row_indices, start=2):
            for c in range(1, self.last_col + 1):
                new_ws.cell(row=idx, column=c, 
                          value=self.ws.cell(row=r, column=c).value)
            new_ws[f"A{idx}"].value = "=ROW()-1"


if __name__ == "__main__":
    test_file = "/Users/smith/Documents/github/OKMart/sabangnet_API/files/test-[기본양식]-합포장용.xlsx"
    ali_merge_packaging(test_file)
    print("모든 처리가 완료되었습니다!")
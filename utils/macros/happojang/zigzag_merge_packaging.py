"""지그재그 합포장 자동화 모듈"""

from __future__ import annotations
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List

from openpyxl.styles import Font, PatternFill, Border
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from utils.excel_handler import ExcelHandler


# 설정 상수
MALL_NAME = "지그재그"
BLUE_FILL = PatternFill(start_color="CCE8FF", end_color="CCE8FF", fill_type="solid")


FONT_MALGUN = Font(name="맑은 고딕", size=9)
HDR_FILL = PatternFill(start_color="006100", end_color="006100", fill_type="solid")
NO_BORDER = Border()

STAR_QTY_RE = re.compile(r"\* ?(\d+)")
MULTI_SEP_RE = re.compile(r"[\/;]")


def to_num(val) -> float:
    try:
        return float(re.sub(r"[^\d.-]", "", str(val))) if str(val).strip() else 0.0
    except ValueError:
        return 0.0


class ZIGZAGDataCleanerUtils:
    @staticmethod
    def clean_product_text(txt: str | None) -> str:
        """
        🔄 ExcelHandler 후보
        상품명 문자열 정리 - ' 1개' 제거
        """
        return str(txt).replace(" 1개", "").strip() if txt else ""

    @staticmethod
    def build_lookup_map(ws_lookup: Worksheet) -> Dict[str, str]:
        """
        Sheet1의 A:B를 딕셔너리로 변환
        (M열 → V열 매핑용 VLOOKUP 대체)
        """
        return {
            str(r[0]): r[1]
            for r in ws_lookup.iter_rows(min_row=2, max_col=2, values_only=True)
            if r[0] is not None
        }


def convert_m_column_to_int(ws: Worksheet) -> None:
    """
    🔄 ExcelHandler 후보
    M열 값을 정수로 변환
    """
    for row in range(2, ws.max_row + 1):
        try:
            cell = ws[f"M{row}"]
            cell.value = int(float(cell.value or 0))
        except (ValueError, TypeError):
            cell.value = 0


def highlight_multiple_items(ws: Worksheet) -> None:
    """
    🔄 ExcelHandler 후보
    F열에서 다중 수량 항목 파란색 배경으로 강조
    """
    for row in range(2, ws.max_row + 1):
        f_cell = ws[f"F{row}"]
        clean_text = ZIGZAGDataCleanerUtils.clean_product_text(f_cell.value)
        f_cell.value = clean_text
        
        # '개' 문자가 2회 이상 등장하면 파란색 배경
        if clean_text.count("개") >= 2:
            f_cell.fill = BLUE_FILL


class ZIGZAGSheetSplitter:
    """시트 분리 처리 클래스"""
    
    def __init__(self, ws: Worksheet):
        self.ws = ws
        self.last_row = ws.max_row
        self.last_col = ws.max_column
        
        # 열 너비 저장
        self.col_widths = [
            ws.column_dimensions[get_column_letter(c)].width
            for c in range(1, self.last_col + 1)
        ]

    def get_rows_by_sheet(self) -> Dict[str, List[int]]:
        """사이트별 행 번호 매핑 생성"""
        site_rows = defaultdict(list)
        for r in range(2, self.last_row + 1):
            text = str(self.ws[f"B{r}"].value or "")
            if "[오케이마트]" in text:
                site_rows["OK"].append(r)
            elif "[아이예스]" in text:
                site_rows["IY"].append(r)
        return site_rows

    def copy_to_new_sheet(self, 
                         wb: Workbook, 
                         sheet_name: str, 
                         row_indices: List[int]) -> None:
        """지정된 행들을 새 시트로 복사"""
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            
        new_ws = wb.create_sheet(sheet_name)
        
        # 헤더와 열 너비 복사
        for c in range(1, self.last_col + 1):
            new_ws.cell(row=1, column=c, 
                       value=self.ws.cell(row=1, column=c).value)
            new_ws.column_dimensions[get_column_letter(c)].width = self.col_widths[c - 1]
        
        # 데이터 복사
        for idx, r in enumerate(row_indices, start=2):
            for c in range(1, self.last_col + 1):
                new_ws.cell(row=idx, column=c, 
                          value=self.ws.cell(row=r, column=c).value)
            new_ws[f"A{idx}"].value = "=ROW()-1"


def zigzag_merge_packaging(input_path: str) -> str:
    """지그재그 주문 합포장 자동화 처리"""
    # Excel 파일 로드
    ex = ExcelHandler.from_file(input_path)
    ws = ex.ws

    # 1. 기본 서식 적용
    ex.set_basic_format()
    
    # 2. M열 정수 변환
    convert_m_column_to_int(ws)
    
    # 3. M열 → V열 VLOOKUP 처리
    if "Sheet1" in ex.wb.sheetnames:
        lookup_map = ZIGZAGDataCleanerUtils.build_lookup_map(ex.wb["Sheet1"])
        for row in range(2, ws.max_row + 1):
            m_val = str(ws[f"M{row}"].value)
            ws[f"V{row}"].value = lookup_map.get(m_val, "")
    
    # 4. D열 수식 설정 (=U+V)
    ex.autofill_d_column(formula="=U{row}+V{row}")
    
    # 5. 상품정보 처리 (다중수량 강조)
    highlight_multiple_items(ws)
    
    # 6. A열 순번 설정
    ex.set_row_number(ws)
    
    # 7. 열 정렬
    ex.set_column_alignment()
    
    # 8. 배경·테두리 제거
    ex.clear_fills_from_second_row()
    ex.clear_borders()
    
    # 9. C→B 정렬
    ex.sort_by_columns([2, 3])
    
    # 10. 시트 분리 (OK, IY)
    splitter = ZIGZAGSheetSplitter(ws)
    rows_by_sheet = splitter.get_rows_by_sheet()
    
    for sheet_name, row_indices in rows_by_sheet.items():
        if row_indices:  # 해당 사이트의 데이터가 있는 경우만
            splitter.copy_to_new_sheet(ex.wb, sheet_name, row_indices)

    # 11. 시트 순서 정리
    desired = ["자동화", "OK", "IY"]
    for name in reversed(desired):
        if name in ex.wb.sheetnames:
            ex.wb._sheets.insert(0, ex.wb._sheets.pop(ex.wb.sheetnames.index(name)))
    
    # 저장
    base_name = Path(input_path).stem  # 확장자 제거한 파일명
    output_path = ex.happojang_save_file(base_name=base_name)
    
    print(f"◼︎ [{MALL_NAME}] 합포장 자동화 완료!")
    
    return output_path


if __name__ == "__main__":
    test_xlsx = "/Users/smith/Documents/github/OKMart/sabangnet_API/files/test-[기본양식]-합포장용.xlsx"
    zigzag_merge_packaging(test_xlsx)
    print("모든 처리가 완료되었습니다!")
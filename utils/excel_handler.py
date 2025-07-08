import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border
import re
import pandas as pd
from openpyxl.utils import get_column_letter

"""
주문관리 Excel 파일 매크로 공통 처리 메소드
- 기본 서식 설정
- 수식 처리
- 데이터 정리
- 정렬 및 레이아웃
- 특수 처리 (제주도 주소, 결제방식 등)
"""


class ExcelHandler:
    def __init__(self, ws, wb=None):
        self.ws = ws
        self.wb = wb
        self.last_row = ws.max_row

    @classmethod
    def from_file(cls, file_path, sheet_index=0):
        """
        파일 경로로 부터 엑셀 파일 로드
        예시:
            ex = ExcelHandler.from_file(file_path)
            ws = ex.ws
            wb = ex.wb
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.worksheets[sheet_index]
        return cls(ws, wb)

    def save_file(self, file_path):
        """
        엑셀 파일 저장
        예시:
            ex.save_file('file.xlsx')
        """
        if file_path.endswith('_매크로_완료.xlsx'):
            output_path = file_path
        else:
            output_path = file_path.replace('.xlsx', '_매크로_완료.xlsx')
        self.wb.save(output_path)
        return output_path

    # 기본 서식 설정 Method
    def set_basic_format(self, header_rgb="006100"):
        """
        폰트, 행높이, 첫 행 배경색, 줄바꿈 해제 등 기본 서식 적용
        예시:
            wb = openpyxl.load_workbook('file.xlsx')
            ws = wb.active
            set_basic_format(ws)
        """
        font = Font(name='맑은 고딕', size=9)
        green_fill = PatternFill(start_color=header_rgb,
                                 end_color=header_rgb, fill_type="solid")
        for row in self.ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = Alignment(wrap_text=False)
            self.ws.row_dimensions[row[0].row].height = 15
        for cell in self.ws[1]:
            cell.fill = green_fill
            cell.alignment = Alignment(horizontal='center')

    # 수식 처리 Method
    def autofill_d_column(self, ws=None, start_row=2, end_row=None, formula=None):
        """
        D열 수식 활성화 및 복사 (금액 계산)
        예시:
            autofill_d_column(ws, 2, last_row, "=U{row}+V{row}")
        - formula에 "{row}"를 포함하면 각 행 번호로 치환하여 적용
        """

        if ws is None:
            ws = self.ws
        if not end_row:
            end_row = self.last_row
        if not formula:
            formula = ws['D2'].value
        for row in range(start_row, end_row + 1):
            if ws[f'D{row}'].value is None:
                continue
            # D열 숫자 포맷 초기화
            ws[f'D{row}'].number_format = 'General'

            # 수식 적용
            if isinstance(formula, str) and '{row}' in formula:
                ws[f'D{row}'].value = formula.format(row=row)
            elif isinstance(formula, str) and '=' in formula:
                ws[f'D{row}'].value = formula.replace('2', str(row))
            else:
                ws[f'D{row}'].value = formula

    def set_row_number(self, ws, start_row=2, end_row=None):
        """
        A열 순번 자동 생성 (=ROW()-1)
        예시:
            set_row_number(ws)
        """
        if not end_row:
            end_row = self.last_row
        if ws is None:
            ws = self.ws
        for row in range(start_row, end_row + 1):
            ws[f'A{row}'].number_format = 'General'
            ws[f"A{row}"].value = "=ROW()-1"

    def convert_formula_to_value(self):
        """
        수식 → 값 변환 처리 (모든 시트)
        (openpyxl은 수식 결과값을 계산하지 않으므로, 실제 값 변환은 Excel에서 복사-값붙여넣기로 처리)
        예시:
            convert_formula_to_value(ws)
        """
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    # 실제 계산은 Excel에서, 여기선 수식 문자열만 제거
                    cell.value = cell.value

    # 데이터 정리 Method

    def clear_borders(self, ws=None):
        """
        테두리 제거
        예시:
            clear_borders(ws)
        """
        if ws is None:
            ws = self.ws
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border()

    def clear_fills_from_second_row(self):
        """
        배경색 제거
        예시:
            clear_fills_from_second_row(ws)
        - 두번째 행부터 모든 셀의 배경색을 제거합니다.
        """
        for row in self.ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = PatternFill(fill_type=None)

    def format_phone_number(self, val):
        """
        전화번호 11자리 → 010-0000-0000 형식
        예시:
            ws['H2'].value = format_phone_number(ws['H2'].value)
        """
        val = str(val or '').replace('-', '').strip()
        if len(val) == 11 and val.startswith('010') and val.isdigit():
            return f"{val[:3]}-{val[3:7]}-{val[7:]}"
        return val

    def clean_model_name(self, val):
        """
        모델명에서 ' 1개' 텍스트 제거
        예시:
            ws['F2'].value = clean_model_name(ws['F2'].value)
        """
        return str(val).replace(' 1개', '') if val else val

    def sum_prow_with_slash(self):
        """
        P열 "/" 금액 합산
        예시:
            sum_prow_with_slash(ws)
        """
        last_row = self.ws.max_row
        for r in range(2, last_row + 1):
            p_raw = str(self.ws[f"P{r}"].value or "")
            if "/" in p_raw:
                nums = [float(n)
                        for n in p_raw.split("/") if n.strip().isdigit()]
                self.ws[f"P{r}"].value = sum(nums) if nums else 0
            else:
                self.ws[f"P{r}"].value = self.to_num(p_raw)

    def to_num(self, val) -> float:
        """
        '12,345원' → 12345.0 (실패 시 0)
        예시:
            num = to_num("12,345원")
        """
        try:
            return float(re.sub(r"[^\d.-]", "", str(val))) if str(val).strip() else 0.0
        except ValueError:
            return 0.0

    def convert_numeric_strings(self, ws=None, start_row: int = 2, end_row: int | None = None, cols: tuple[str, ...] | None = None) -> None:
        """
        워크시트의 문자열 숫자를 숫자 타입(int/float)으로 변환합니다.

        Args:
            start_row: 변환을 시작할 데이터 행 (헤더가 1행이라고 가정).
            end_row: 변환을 끝낼 행. None이면 마지막 행까지 처리.
            cols: 변환할 열 머리글(예: ("E","M","Q","W")). None이면 모든 열을 대상.

        사용 예시:
            # 1) 모든 열 대상
            ex.convert_numeric_strings()

            # 2) 특정 열만
            ex.convert_numeric_strings(cols=("E", "M", "Q", "W"))
        """
        if ws is None:
            ws = self.ws
        if end_row is None:
            end_row = self.ws.max_row

        # 변환 대상 열 결정
        if cols:
            target_cols = cols
        else:
            # 1행 헤더를 기준으로 모든 실제 열 레터를 수집
            target_cols = tuple(
                cell.column_letter for cell in ws[1] if cell.value is not None)

        for row in range(start_row, end_row + 1):
            for col in target_cols:
                cell = ws[f"{col}{row}"]
                if cell.value is None:
                    continue
                if isinstance(cell.value, str):
                    raw = cell.value.strip()
                    # 숫자(0-9), 쉼표, 마침표 외 다른 문자가 섞여 있으면 변환하지 않음
                    if re.fullmatch(r"[0-9,\.]+", raw):
                        num_val = self.to_num(raw)
                        # 0 도 유효 숫자로 인정
                        if raw not in {"", ".", ","}:
                            cell.value = num_val
                            cell.number_format = "General"

    # 정렬 및 레이아웃 Method

    def set_column_alignment(self, ws=None):
        """
        A,B(가운데), D,E,G(오른쪽), 첫 행 가운데 정렬
        예시:
            set_column_alignment(ws)
        """
        center = Alignment(horizontal='center')
        right = Alignment(horizontal='right')

        align_map = {
            'center': {'A', 'B'},
            'right': {'D', 'E', 'G'}
        }
        if ws is None:
            ws = self.ws
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                col_letter = cell.column_letter
                if col_letter in align_map['center']:
                    cell.alignment = center
                elif col_letter in align_map['right']:
                    cell.alignment = right

        # 1행: 모든 셀 가운데 정렬
        for cell in self.ws[1]:
            cell.alignment = center

    def sort_dataframe_by_c_b(self, df, c_col='C', b_col='B'):
        """
        DataFrame을 C열 → B열 순서로 오름차순 정렬
        예시:
            df = sort_dataframe_by_c_b(df)
        """
        if c_col in df.columns and b_col in df.columns:
            print(c_col, b_col)
            return df.sort_values(by=[c_col, b_col]).reset_index(drop=True)
        return df

    # 특수 처리 Method

    def process_jeju_address(self, row, f_col='F', j_col='J'):
        """
        제주도 주소: '[3000원 연락해야함]' 추가, 연한 파란색 배경 및 빨간 글씨 적용
        예시:
            process_jeju_address(ws, row=5)
        """
        red_font = Font(color="FF0000", bold=True)
        # RGB(204,255,255) → hex: "CCFFFF"
        light_blue_fill = PatternFill(
            start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
        # F열 안내문 추가
        f_val = self.ws[f'{f_col}{row}'].value
        if f_val and "[3000원 연락해야함]" not in str(f_val):
            self.ws[f'{f_col}{row}'].value = str(f_val) + " [3000원 연락해야함]"
        # J열 빨간 글씨
        self.ws[f'{j_col}{row}'].font = red_font
        # F열 연한 파란색 배경
        self.ws[f'{f_col}{row}'].fill = light_blue_fill

    def process_l_column(self, row, l_col='L'):
        """
        L열 결제방식: '신용' 삭제, '착불' 빨간 글씨
        예시:
            process_l_column(ws, row=7)
        """
        red_font = Font(color="FF0000", bold=True)
        l_val = self.ws[f'{l_col}{row}'].value
        if l_val == "신용":
            self.ws[f'{l_col}{row}'].value = ""
        elif l_val == "착불":
            self.ws[f'{l_col}{row}'].font = red_font

    def highlight_column(self, col: str, light_color: PatternFill, ws=None, start_row: int = 2, last_row: int = None):
        """
        특정 열 하이라이트 처리
        예시:
            - F열 모르겠는 셀 색칠음영 (하늘색)
            - highlight_column(col='F', light_color=light_blue_fill, start_row=2, last_row=last_row)
        """

        def _should_highlight(txt: str) -> bool:
            """
            셀 값이 다음 조건 중 하나라도 만족하면 True:
            - 빈 문자열
            - 'none' (대소문자 무관)
            - 모든 문자가 '#'
            - 순수 숫자
            - '숫자개' 패턴 (예: '3개')
            """
            if not txt or txt.strip() == "":
                return True
            txt = txt.strip()
            if txt.lower() == "none":
                return True
            if all(c == '#' for c in txt):
                return True
            if txt.isdigit():
                return True
            if txt.endswith("개") and txt[:-1].isdigit():
                return True
            return False

        if ws is None:
            ws = self.ws
        if not last_row:
            last_row = self.last_row
        for row in range(start_row, last_row + 1):
            cell_value = ws[f'{col}{row}'].value
            txt = str(cell_value).strip() if cell_value else ""

            if cell_value is not None and _should_highlight(txt):
                ws[f"{col}{row}"].fill = light_color

    def set_header_style(self, ws, headers: list, fill: PatternFill, font: Font, alignment: Alignment):
        """
        지정한 워크시트의 헤더 행에 배경색, 폰트, 정렬을 일괄 적용
        Args:
            ws: openpyxl worksheet
            headers: 헤더 리스트
            fill: 헤더 배경색(hex)
            font: 폰트
            alignment: 정렬 
        """
        for col in range(1, len(headers) + 1):
            header_value = headers[col-1]
            ws_cell = ws.cell(row=1, column=col, value=header_value)
            ws_cell.fill = fill
            ws_cell.font = font
            ws_cell.alignment = alignment

    def convert_to_number(self, cell_value):
        """
        문자열을 숫자로 변환
        예시:
            convert_to_number(ws['M2'].value)
        """
        return float(cell_value) if '.' in str(cell_value) else int(float(cell_value))

    def to_dataframe(self, ws=None, start_row=2, start_col=1, end_row=None, end_col=None):
        """
        지정된 워크시트의 데이터를 DataFrame으로 변환
        args:
            ws: 워크시트
            start_row: 시작 행
            start_col: 시작 열
            end_row: 끝 행
            end_col: 끝 열
        """
        ws = ws or self.ws
        end_row = end_row or ws.max_row
        end_col = end_col or ws.max_column

        # 헤더 추출
        headers = []

        for col in range(start_col, end_col + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(header if header else f"Col{col}")

        # 데이터 추출
        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                row_data.append(ws.cell(row=row, column=col).value)
            data.append(row_data)

        return pd.DataFrame(data, columns=headers)

    def create_split_sheets(self, headers: list, sheet_names: list):
        """
        지정한 이름의 시트를 생성하고, 열 너비/행 높이만 원본 시트(self.ws)에서 복사합니다.
        헤더 복사 및 스타일 적용은 제외합니다.

        Args:
            headers (list): 헤더 리스트 
            sheet_names (list): 생성할 시트명 리스트 ["OK,CL,BB", "IY"]

        Returns:
            dict: {시트명: 워크시트 객체}
        """
        ws_map = {}
        for sheet_name in sheet_names:
            # 기존 시트 삭제
            if sheet_name in self.wb.sheetnames:
                del self.wb[sheet_name]
            # 새 시트 생성
            ws = self.wb.create_sheet(title=sheet_name)
            # 열 너비 복사
            for col in range(1, len(headers) + 1):
                col_letter = get_column_letter(col)
                src_width = self.ws.column_dimensions[col_letter].width
                ws.column_dimensions[col_letter].width = src_width
            # 행 높이 복사 (헤더 행만)
            ws.row_dimensions[1].height = 15
            ws_map[sheet_name] = ws
        return ws_map
